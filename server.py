#!/usr/bin/env python3
"""
Posgrado CRM Enterprise v3.0 - Servidor de Desarrollo (Python)
===============================================================
- CRUD completo: Programas, Usuarios, Enlaces, Inscripciones
- Pipeline de estados para inscripciones
- Generacion de enlaces rastreables por asesor+programa
- Busqueda y filtrado en tablas
- Analytics con barras de progreso y metricas
- Protecciones de integridad y validaciones
"""

import os
import sys
import json
import http.server
import urllib.parse
from http.cookies import SimpleCookie
import random
import hashlib
from datetime import datetime, timedelta
import time

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')
if hasattr(sys.stderr, 'reconfigure'):
    sys.stderr.reconfigure(encoding='utf-8')

PORT = 3000

# ============================================================
# BASE DE DATOS EN MEMORIA
# ============================================================
db = {
    "sessions": {},
    "users": [
        {"id": "u1", "name": "Administrador Principal", "email": "admin@posgrado.com", "password": "Admin123!", "role": "ADMIN", "phone": "+591 70000000", "active": True, "createdAt": "2026-01-15 08:00"},
        {"id": "u2", "name": "Juan Perez", "email": "juan.perez@posgrado.com", "password": "Asesor123!", "role": "ASESOR", "phone": "+591 71111111", "active": True, "createdAt": "2026-03-01 09:00"},
        {"id": "u3", "name": "Maria Lopez", "email": "maria.lopez@posgrado.com", "password": "Asesor123!", "role": "ASESOR", "phone": "+591 72222222", "active": True, "createdAt": "2026-03-15 09:00"},
        {"id": "u4", "name": "Carlos Gomez", "email": "carlos.gomez@posgrado.com", "password": "Asesor123!", "role": "ASESOR", "phone": "+591 73333333", "active": True, "createdAt": "2026-04-01 09:00"},
    ],
    "programs": [
        {"id": "p0", "name": "Curso Especializado en Marketing Digital e IA", "code": "CMD-2026", "type": "CURSO", "active": True, "description": "Programa intensivo en estrategias de marketing digital potenciadas por inteligencia artificial.", "duration": "3 meses", "modality": "Virtual", "investment": "2.500 Bs", "imageUrl": "/uploads/marketing_ia.png"},
        {"id": "p1", "name": "Maestria en Educacion Superior y Docencia Universitaria", "code": "MED-2026", "type": "MAESTRIA", "active": True, "description": "Formacion avanzada en pedagogia universitaria y gestion educativa.", "duration": "24 meses", "modality": "Semipresencial", "investment": "18.000 Bs", "imageUrl": "/uploads/educacion_superior.png"},
        {"id": "p2", "name": "Maestria en Derecho Corporativo y Financiero", "code": "MDC-2026", "type": "MAESTRIA", "active": True, "description": "Especializacion en derecho empresarial y regulacion financiera.", "duration": "24 meses", "modality": "Presencial", "investment": "20.000 Bs", "imageUrl": "https://images.unsplash.com/photo-1450133064473-71024230f91b?w=800&auto=format&fit=crop&q=80"},
        {"id": "p3", "name": "Especialidad en Salud Publica y Epidemiologia", "code": "ESP-2026", "type": "ESPECIALIDAD", "active": True, "description": "Formacion en gestion sanitaria y analisis epidemiologico.", "duration": "12 meses", "modality": "Semipresencial", "investment": "12.000 Bs", "imageUrl": "https://images.unsplash.com/photo-1576091160399-112ba8d25d1d?w=800&auto=format&fit=crop&q=80"},
        {"id": "p4", "name": "Diplomado en Finanzas Corporativas y Fintech", "code": "DFC-2026", "type": "DIPLOMADO", "active": True, "description": "Programa en finanzas empresariales y tecnologias financieras.", "duration": "6 meses", "modality": "Virtual", "investment": "5.000 Bs", "imageUrl": "https://images.unsplash.com/photo-1559526324-4b87b5e36e44?w=800&auto=format&fit=crop&q=80"},
    ],
    "links": [
        {"id": "l1", "code": "4fd89af8b2", "programId": "p1", "advisorId": "u2", "clickCount": 14, "active": True, "createdAt": "2026-06-15 10:00"},
        {"id": "l2", "code": "8ab12c34de", "programId": "p2", "advisorId": "u2", "clickCount": 9, "active": True, "createdAt": "2026-06-15 10:05"},
        {"id": "l3", "code": "99x77y66z1", "programId": "p3", "advisorId": "u3", "clickCount": 22, "active": True, "createdAt": "2026-06-20 14:00"},
        {"id": "l4", "code": "33aa44bb55", "programId": "p4", "advisorId": "u4", "clickCount": 18, "active": True, "createdAt": "2026-06-22 11:00"},
        {"id": "l5", "code": "abc123def4", "programId": "p0", "advisorId": "u2", "clickCount": 7, "active": True, "createdAt": "2026-07-01 09:00"},
        {"id": "l6", "code": "ff11ee22dd", "programId": "p1", "advisorId": "u3", "clickCount": 5, "active": True, "createdAt": "2026-07-05 16:00"},
    ],
    "registrations": [
        {"id": "r1", "firstNames": "Roberto", "lastNames": "Vargas Morales", "fullName": "Roberto Vargas Morales", "ci": "7654321 LP", "email": "roberto.vargas@gmail.com", "phone": "77889900", "address": "Av. 6 de Agosto #2450, Sopocachi", "city": "La Paz", "birthDate": "1992-05-14", "civilStatus": "Soltero/a", "ciFrontUrl": "/uploads/marketing_ia.png", "ciBackUrl": "/uploads/educacion_superior.png", "academicDegree": "Licenciatura", "profession": "Ingeniero Comercial", "university": "UMSA", "programId": "p1", "advisorId": "u2", "linkId": "l1", "status": "CONTACTADO", "createdAt": "2026-07-28 10:15", "notes": "Interesado en becas parciales"},
        {"id": "r2", "firstNames": "Ana Isabel", "lastNames": "Mendoza", "fullName": "Ana Isabel Mendoza", "ci": "8877665 CB", "email": "ana.mendoza@hotmail.com", "phone": "71234567", "address": "Calle España #450, Centro", "city": "Cochabamba", "birthDate": "1994-08-22", "civilStatus": "Casado/a", "ciFrontUrl": "/uploads/educacion_superior.png", "ciBackUrl": "/uploads/marketing_ia.png", "academicDegree": "Licenciatura", "profession": "Licenciada en Educación", "university": "UMSS", "programId": "p1", "advisorId": "u2", "linkId": "l1", "status": "NUEVO", "createdAt": "2026-07-28 11:30", "notes": ""},
        {"id": "r3", "firstNames": "Fernando", "lastNames": "Aguilera Rivas", "fullName": "Dr. Fernando Aguilera Rivas", "ci": "4567890 SC", "email": "f.aguilera@hospital.org", "phone": "76543210", "address": "Av. San Martín #800, Equipetrol", "city": "Santa Cruz", "birthDate": "1988-11-03", "civilStatus": "Casado/a", "ciFrontUrl": "/uploads/marketing_ia.png", "ciBackUrl": "/uploads/educacion_superior.png", "academicDegree": "Especialidad", "profession": "Médico Cirujano", "university": "UAGRM", "programId": "p3", "advisorId": "u3", "linkId": "l3", "status": "MATRICULADO", "createdAt": "2026-07-27 14:20", "notes": "Documentacion completa. Pago realizado."},
        {"id": "r4", "firstNames": "Claudia Patricia", "lastNames": "Suarez", "fullName": "Claudia Patricia Suarez", "ci": "6543210 LP", "email": "claudia.suarez@banco.com", "phone": "70987654", "address": "Calle Calacoto 18 #300", "city": "La Paz", "birthDate": "1995-02-18", "civilStatus": "Soltero/a", "ciFrontUrl": "/uploads/educacion_superior.png", "ciBackUrl": "/uploads/marketing_ia.png", "academicDegree": "Licenciatura", "profession": "Auditora / Financiera", "university": "UCB", "programId": "p4", "advisorId": "u4", "linkId": "l4", "status": "DOC_PENDIENTE", "createdAt": "2026-07-26 16:45", "notes": "Falta titulo de licenciatura"},
        {"id": "r5", "firstNames": "Marco Antonio", "lastNames": "Rios", "fullName": "Marco Antonio Rios", "ci": "5432198 SC", "email": "marco.rios@empresa.com", "phone": "77001122", "address": "Barrio Sirari, Calle 4 #12", "city": "Santa Cruz", "birthDate": "1990-09-12", "civilStatus": "Conviviente", "ciFrontUrl": "/uploads/marketing_ia.png", "ciBackUrl": "/uploads/educacion_superior.png", "academicDegree": "Licenciatura", "profession": "Abogado Corporativo", "university": "UPSA", "programId": "p2", "advisorId": "u2", "linkId": "l2", "status": "COMPLETO", "createdAt": "2026-07-25 09:00", "notes": "Documentos verificados"},
        {"id": "r6", "firstNames": "Lucia Fernanda", "lastNames": "Gutierrez", "fullName": "Lucia Fernanda Gutierrez", "ci": "3345678 OR", "email": "lucia.gutierrez@uni.edu", "phone": "79988776", "address": "Av. 6 de Octubre #112", "city": "Oruro", "birthDate": "1997-04-05", "civilStatus": "Soltero/a", "ciFrontUrl": "/uploads/educacion_superior.png", "ciBackUrl": "/uploads/marketing_ia.png", "academicDegree": "Licenciatura", "profession": "Comunicadora Digital", "university": "UTO", "programId": "p0", "advisorId": "u2", "linkId": "l5", "status": "NUEVO", "createdAt": "2026-07-29 08:10", "notes": ""},
        {"id": "r7", "firstNames": "Pedro", "lastNames": "Saavedra", "fullName": "Ing. Pedro Saavedra", "ci": "1122334 CH", "email": "p.saavedra@tech.com", "phone": "76655443", "address": "Calle Junín #55", "city": "Sucre", "birthDate": "1991-07-30", "civilStatus": "Casado/a", "ciFrontUrl": "/uploads/marketing_ia.png", "ciBackUrl": "/uploads/educacion_superior.png", "academicDegree": "Maestría", "profession": "Ingeniero de Sistemas", "university": "USFX", "programId": "p3", "advisorId": "u3", "linkId": "l3", "status": "CONTACTADO", "createdAt": "2026-07-29 09:30", "notes": "Consulto por modalidad"},
        {"id": "r8", "firstNames": "Daniela", "lastNames": "Rojas Quiroga", "fullName": "Daniela Rojas Quiroga", "ci": "9988776 TJ", "email": "daniela.rojas@correo.com", "phone": "71122334", "address": "Calle La Madrid #400", "city": "Tarija", "birthDate": "1996-12-01", "civilStatus": "Soltero/a", "ciFrontUrl": "/uploads/educacion_superior.png", "ciBackUrl": "/uploads/marketing_ia.png", "academicDegree": "Licenciatura", "profession": "Psicóloga", "university": "UAJMS", "programId": "p1", "advisorId": "u3", "linkId": "l6", "status": "DOC_PENDIENTE", "createdAt": "2026-07-24 15:00", "notes": "Envio parcial de documentos"},
    ],
    "form_settings": {
        "datos_personales": True,
        "documentos_ci": True,
        "datos_contacto": True,
        "datos_academicos": True
    }
}


# ============================================================
# HELPERS
# ============================================================
STATUS_PIPELINE = ["NUEVO", "CONTACTADO", "DOC_PENDIENTE", "COMPLETO", "MATRICULADO"]
STATUS_LABELS = {"NUEVO": "Nuevo", "CONTACTADO": "Contactado", "DOC_PENDIENTE": "Doc. Pendiente", "COMPLETO": "Completo", "MATRICULADO": "Matriculado"}
STATUS_COLORS = {
    "NUEVO": ("bg-sky-500/15 text-sky-400 border-sky-500/30", "sky"),
    "CONTACTADO": ("bg-indigo-500/15 text-indigo-400 border-indigo-500/30", "indigo"),
    "DOC_PENDIENTE": ("bg-amber-500/15 text-amber-400 border-amber-500/30", "amber"),
    "COMPLETO": ("bg-emerald-500/15 text-emerald-400 border-emerald-500/30", "emerald"),
    "MATRICULADO": ("bg-violet-500/15 text-violet-400 border-violet-500/30", "violet"),
}
TYPE_COLORS = {
    "MAESTRIA": "bg-violet-500/15 text-violet-400 border-violet-500/30",
    "ESPECIALIDAD": "bg-cyan-500/15 text-cyan-400 border-cyan-500/30",
    "DIPLOMADO": "bg-amber-500/15 text-amber-400 border-amber-500/30",
    "CURSO": "bg-blue-500/15 text-blue-400 border-blue-500/30",
}

MONTH_NAMES = {
    "01": "Enero", "02": "Febrero", "03": "Marzo", "04": "Abril",
    "05": "Mayo", "06": "Junio", "07": "Julio", "08": "Agosto",
    "09": "Septiembre", "10": "Octubre", "11": "Noviembre", "12": "Diciembre"
}

def format_ym_label(ym):
    if not ym or len(ym) < 7: return str(ym)
    y, m = ym[:4], ym[5:7]
    return f"{MONTH_NAMES.get(m, m)} {y}"

def get_all_12_months_options(selected_ym="all", regs=None):
    curr_now = datetime.now()
    curr_year = curr_now.year
    curr_ym = curr_now.strftime("%Y-%m")

    years_set = {curr_year}
    if regs:
        for r in regs:
            dt = r.get("createdAt", "")
            if len(dt) >= 4 and dt[:4].isdigit():
                years_set.add(int(dt[:4]))

    sorted_years = sorted(list(years_set), reverse=True)

    options_html = ""
    for yr in sorted_years:
        options_html += f'<optgroup label="Año {yr}">'
        for m in range(1, 13):
            ym_val = f"{yr}-{m:02d}"
            sel = "selected" if selected_ym == ym_val else ""
            lbl = format_ym_label(ym_val)
            tag = " ⭐ (MES ACTUAL)" if ym_val == curr_ym else ""
            
            m_count = len([r for r in (regs or []) if r.get("createdAt","").startswith(ym_val)])
            count_tag = f" — [{m_count} contactos]" if m_count > 0 else ""
            
            options_html += f'<option value="{ym_val}" {sel}>{lbl}{tag}{count_tag}</option>'
        options_html += '</optgroup>'
    
    return options_html

def get_user(uid): return next((u for u in db["users"] if u["id"] == uid), None)
def get_program(pid): return next((p for p in db["programs"] if p["id"] == pid), None)
def get_link(lid): return next((l for l in db["links"] if l["id"] == lid), None)
def get_link_by_code(code): return next((l for l in db["links"] if l["code"] == code), None)

def get_uid_from_cookie(cookie_str):
    c = SimpleCookie()
    c.load(cookie_str or "")
    t = c.get("crm_session")
    return db["sessions"].get(t.value) if t else None

def gen_id(prefix): return f"{prefix}_{random.randint(10000, 99999)}"
def gen_code(): return hashlib.md5(str(random.random()).encode()).hexdigest()[:10]

def program_name(pid):
    p = get_program(pid)
    return p["name"] if p else "Programa eliminado"

def program_code(pid):
    p = get_program(pid)
    return p["code"] if p else "N/A"

def user_name(uid):
    u = get_user(uid)
    return u["name"] if u else "Usuario eliminado"

def regs_for_link(lid): return [r for r in db["registrations"] if r.get("linkId") == lid]
def regs_for_program(pid): return [r for r in db["registrations"] if r["programId"] == pid]
def regs_for_advisor(uid): return [r for r in db["registrations"] if r["advisorId"] == uid]
def links_for_program(pid): return [l for l in db["links"] if l["programId"] == pid]
def links_for_advisor(uid): return [l for l in db["links"] if l["advisorId"] == uid]

def sync_advisor_links():
    all_users = [u for u in db["users"] if u.get("active")]
    active_programs = [p for p in db["programs"] if p.get("active")]
    created = 0
    for u in all_users:
        for prog in active_programs:
            existing = next((l for l in db["links"] if l["programId"] == prog["id"] and l["advisorId"] == u["id"]), None)
            if not existing:
                db["links"].append({
                    "id": gen_id("l"),
                    "code": gen_code(),
                    "programId": prog["id"],
                    "advisorId": u["id"],
                    "clickCount": 0,
                    "active": True,
                    "createdAt": datetime.now().strftime("%Y-%m-%d %H:%M")
                })
                created += 1
    return created

# Sincronizar enlaces iniciales de la base de datos
sync_advisor_links()

def esc(s):
    """HTML-escape a string to prevent injection in form values."""
    return str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace('"', "&quot;")


def generate_excel_bytes(regs, is_admin, selected_month, selected_advisor, host_url="http://localhost:3000"):
    """Genera un archivo Excel (.xlsx) nativo con formato profesional y enlaces directos a las fotos de C.I."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Contactos Posgrado"

    # Estilos profesionales de Excel
    title_font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    title_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")

    meta_font = Font(name="Calibri", size=10, italic=True, color="64748B")

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")

    zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    thin_border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1")
    )

    link_font = Font(name="Calibri", size=10, color="2563EB", underline="single")

    # Banner Superior de Titulo
    ws.merge_cells("A1:U1")
    ws["A1"] = "POSGRADO CRM ENTERPRISE — REPORTE OFICIAL DE INSCRIPCIONES Y CONTACTOS"
    ws["A1"].font = title_font
    ws["A1"].fill = title_fill
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 32

    # Metadatos del Reporte
    now_str = datetime.now().strftime("%d/%m/%Y %H:%M")
    m_label = format_ym_label(selected_month) if selected_month != "all" else "Todos los meses"
    a_label = user_name(selected_advisor) if (is_admin and selected_advisor != "all") else "Todos los asesores"
    ws.merge_cells("A2:U2")
    ws["A2"] = f"Fecha de Exportacion: {now_str} | Filtro Mes: {m_label} | Filtro Asesor: {a_label} | Total Contactos: {len(regs)}"
    ws["A2"].font = meta_font
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 20

    # Encabezados de Columna
    headers = [
        "Nº", "Fecha Registro", "Nombres", "Apellidos", "Nombre Completo", 
        "C.I. y Extension", "Correo Electronico", "Celular / WhatsApp", "Direccion", "Ciudad", 
        "Fecha Nacimiento", "Estado Civil", "Grado Academico", "Profesion", "Universidad de Egreso", 
        "Foto C.I. Anverso", "Foto C.I. Reverso", "Programa Academico", "Asesor Comercial", 
        "Estado Comercial", "Notas de Seguimiento"
    ]

    ws.append([])  # Fila 3 vacia
    ws.append(headers)  # Fila 4 Encabezados
    ws.row_dimensions[4].height = 26

    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=4, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # Filas de Datos
    for idx, r in enumerate(regs, 1):
        row_num = idx + 4
        c1 = r.get("ciFrontUrl", "")
        c2 = r.get("ciBackUrl", "")

        pname = program_name(r.get("programId"))
        aname = user_name(r.get("advisorId"))
        st_label = STATUS_LABELS.get(r.get("status", ""), r.get("status", ""))

        row_data = [
            idx,
            r.get("createdAt", ""),
            r.get("firstNames", ""),
            r.get("lastNames", ""),
            r.get("fullName", ""),
            r.get("ci", ""),
            r.get("email", ""),
            r.get("phone", ""),
            r.get("address", ""),
            r.get("city", ""),
            r.get("birthDate", ""),
            r.get("civilStatus", ""),
            r.get("academicDegree", ""),
            r.get("profession", ""),
            r.get("university", ""),
            "Ver Foto Anverso" if c1 else "Sin foto",
            "Ver Foto Reverso" if c2 else "Sin foto",
            pname,
            aname,
            st_label,
            r.get("notes", "")
        ]

        ws.append(row_data)
        ws.row_dimensions[row_num].height = 22
        fill_to_use = zebra_fill if idx % 2 == 0 else white_fill

        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.fill = fill_to_use
            cell.border = thin_border
            cell.font = Font(name="Calibri", size=10)
            cell.alignment = Alignment(vertical="center")

            # Enlaces hipervinculados para las fotos de C.I.
            if col_num == 16 and c1:
                url = c1 if c1.startswith("http") else (host_url + c1)
                cell.hyperlink = url
                cell.font = link_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_num == 17 and c2:
                url = c2 if c2.startswith("http") else (host_url + c2)
                cell.hyperlink = url
                cell.font = link_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_num in [1, 2, 6, 8, 10, 11, 20]:
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # Ajuste automatico de ancho de columnas
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row in [1, 2]: continue
            val_str = str(cell.value or "")
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 11)

    import io
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


# ============================================================
# SHARED HTML / CSS
# ============================================================
BASE_CSS = """
* { font-family: 'Inter', system-ui, sans-serif; margin: 0; padding: 0; box-sizing: border-box; }
body { background: #020617; color: #e2e8f0; }
::-webkit-scrollbar { width: 4px; height: 4px; }
::-webkit-scrollbar-track { background: transparent; }
::-webkit-scrollbar-thumb { background: #1e293b; border-radius: 2px; }
::-webkit-scrollbar-thumb:hover { background: #334155; }
a { color: inherit; }

.fade-in { animation: fadeIn 0.2s ease-out; }
@keyframes fadeIn { from { opacity:0; transform:translateY(6px); } to { opacity:1; transform:translateY(0); } }

.card { background: #0f172a; border: 1px solid #1e293b; border-radius: 14px; transition: all 0.2s ease; }
.card:hover { border-color: #334155; }
.card-lift:hover { transform: translateY(-1px); box-shadow: 0 8px 24px rgba(0,0,0,0.25); }

.btn { display: inline-flex; align-items: center; gap: 5px; padding: 7px 14px; border-radius: 8px; font-size: 11px; font-weight: 600; border: none; cursor: pointer; transition: all 0.15s ease; text-decoration: none; line-height: 1.2; }
.btn-primary { background: linear-gradient(135deg, #2563eb, #4f46e5); color: white; box-shadow: 0 2px 8px rgba(37,99,235,0.2); }
.btn-primary:hover { background: linear-gradient(135deg, #3b82f6, #6366f1); box-shadow: 0 4px 12px rgba(37,99,235,0.3); }
.btn-secondary { background: #1e293b; color: #94a3b8; border: 1px solid #1e293b; }
.btn-secondary:hover { background: #334155; color: #e2e8f0; border-color: #334155; }
.btn-danger { background: rgba(239,68,68,0.08); color: #f87171; border: 1px solid rgba(239,68,68,0.15); }
.btn-danger:hover { background: rgba(239,68,68,0.16); }
.btn-success { background: rgba(16,185,129,0.08); color: #34d399; border: 1px solid rgba(16,185,129,0.15); }
.btn-success:hover { background: rgba(16,185,129,0.16); }
.btn-warning { background: rgba(245,158,11,0.08); color: #fbbf24; border: 1px solid rgba(245,158,11,0.15); }
.btn-warning:hover { background: rgba(245,158,11,0.16); }
.btn-sm { padding: 4px 9px; font-size: 10px; border-radius: 6px; }

.input { width: 100%; padding: 9px 12px; border-radius: 8px; background: #020617; border: 1px solid #1e293b; color: white; font-size: 12px; outline: none; transition: all 0.2s; }
.input:focus { border-color: #3b82f6; box-shadow: 0 0 0 2px rgba(59,130,246,0.08); }
.input::placeholder { color: #334155; }
select.input { appearance: none; background-image: url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='10' height='10' viewBox='0 0 24 24' fill='none' stroke='%2364748b' stroke-width='2'%3E%3Cpath d='M6 9l6 6 6-6'/%3E%3C/svg%3E"); background-repeat: no-repeat; background-position: right 10px center; padding-right: 28px; }

.badge { display: inline-flex; align-items: center; padding: 2px 8px; border-radius: 100px; font-size: 9px; font-weight: 700; border: 1px solid; text-transform: uppercase; letter-spacing: 0.4px; }

table { width: 100%; border-collapse: collapse; }
thead { background: rgba(15,23,42,0.6); }
th { padding: 10px 14px; font-size: 9px; font-weight: 700; text-transform: uppercase; letter-spacing: 0.8px; color: #475569; text-align: left; border-bottom: 1px solid #1e293b; }
td { padding: 10px 14px; font-size: 11px; border-bottom: 1px solid rgba(30,41,59,0.4); }
tbody tr { transition: background 0.1s; }
tbody tr:hover { background: rgba(30,41,59,0.25); }

.toast { position:fixed; top:16px; right:16px; z-index:200; animation: slideInRight 0.25s ease-out; }
@keyframes slideInRight { from { opacity:0; transform:translateX(60px); } to { opacity:1; transform:translateX(0); } }

.progress-bar { height: 5px; border-radius: 3px; background: #1e293b; overflow: hidden; }
.progress-fill { height: 100%; border-radius: 3px; transition: width 0.5s ease; }

.public-form-grid { display: grid; grid-template-columns: 1fr 1fr; gap: 32px; align-items: start; max-width: 1180px; width: 100%; margin: 0 auto; }
.public-form-poster { position: sticky; top: 24px; }
.form-grid-2col { display: grid; grid-template-columns: 1fr 1fr; gap: 10px; }
.form-grid-2to1 { display: grid; grid-template-columns: 2fr 1fr; gap: 10px; }
.form-grid-3col { display: grid; grid-template-columns: 1fr 1fr 1fr; gap: 10px; }

@media (max-width: 960px) {
  .public-form-grid { grid-template-columns: 1fr; gap: 20px; }
  .public-form-poster { position: static; }
}

@media (max-width: 640px) {
  .public-form-wrapper { padding: 20px 12px !important; }
  .public-form-header { padding: 4px 6px !important; margin-bottom: 0px !important; }
  .public-form-header h1 { font-size: 21px !important; line-height: 1.25 !important; }
  .public-form-header p { font-size: 11px !important; }
  .public-form-card { padding: 18px 14px !important; border-radius: 16px !important; }
  .form-grid-2col, .form-grid-2to1, .form-grid-3col { grid-template-columns: 1fr !important; gap: 10px !important; }
  .input { font-size: 13px !important; padding: 10px 12px !important; min-height: 42px !important; }
  .btn-primary { min-height: 46px !important; font-size: 13px !important; }
}

/* Sidebar nav */
.nav-item { display:flex; align-items:center; gap:8px; padding:7px 10px; border-radius:7px; font-size:11px; font-weight:500; text-decoration:none; color:#64748b; transition:all 0.12s ease; border:1px solid transparent; }
.nav-item:hover { color:#94a3b8; background:rgba(30,41,59,0.4); }
.nav-item.active { color:#60a5fa; background:rgba(59,130,246,0.08); border-color:rgba(59,130,246,0.12); font-weight:600; }
.nav-item svg { width:15px; height:15px; flex-shrink:0; opacity:0.7; }
.nav-item.active svg { opacity:1; }
.nav-section { font-size:9px; font-weight:700; color:#334155; text-transform:uppercase; letter-spacing:1.2px; padding:12px 10px 4px; }
"""

ICON = {
    "dashboard": '<svg fill="none" stroke="currentColor" stroke-width="1.8" viewBox="0 0 24 24"><rect x="3" y="3" width="7" height="7" rx="1.5"/><rect x="14" y="3" width="7" height="7" rx="1.5"/><rect x="3" y="14" width="7" height="7" rx="1.5"/><rect x="14" y="14" width="7" height="7" rx="1.5"/></svg>',
    "registrations": '<svg fill="none" stroke="currentColor" stroke-width="1.8" viewBox="0 0 24 24"><path d="M9 12h6m-6 4h6m2 5H7a2 2 0 01-2-2V5a2 2 0 012-2h5.586a1 1 0 01.707.293l5.414 5.414a1 1 0 01.293.707V19a2 2 0 01-2 2z"/></svg>',
    "links": '<svg fill="none" stroke="currentColor" stroke-width="1.8" viewBox="0 0 24 24"><path d="M13.828 10.172a4 4 0 00-5.656 0l-4 4a4 4 0 105.656 5.656l1.102-1.101m-.758-4.899a4 4 0 005.656 0l4-4a4 4 0 00-5.656-5.656l-1.1 1.1"/></svg>',
    "programs": '<svg fill="none" stroke="currentColor" stroke-width="1.8" viewBox="0 0 24 24"><path d="M12 6.253v13m0-13C10.832 5.477 9.246 5 7.5 5S4.168 5.477 3 6.253v13C4.168 18.477 5.754 18 7.5 18s3.332.477 4.5 1.253m0-13C13.168 5.477 14.754 5 16.5 5c1.747 0 3.332.477 4.5 1.253v13C19.832 18.477 18.247 18 16.5 18c-1.746 0-3.332.477-4.5 1.253"/></svg>',
    "users": '<svg fill="none" stroke="currentColor" stroke-width="1.8" viewBox="0 0 24 24"><path d="M17 21v-2a4 4 0 00-4-4H5a4 4 0 00-4 4v2"/><circle cx="9" cy="7" r="4"/><path d="M23 21v-2a4 4 0 00-3-3.87m-4-12a4 4 0 010 7.75"/></svg>',
    "plus": '<svg fill="none" stroke="currentColor" stroke-width="2" viewBox="0 0 24 24"><path d="M12 5v14m-7-7h14"/></svg>',
    "back": '<svg fill="none" stroke="currentColor" stroke-width="2" viewBox="0 0 24 24"><path d="M15 19l-7-7 7-7"/></svg>',
    "edit": '<svg fill="none" stroke="currentColor" stroke-width="1.8" viewBox="0 0 24 24"><path d="M11 4H4a2 2 0 00-2 2v14a2 2 0 002 2h14a2 2 0 002-2v-7"/><path d="M18.5 2.5a2.121 2.121 0 013 3L12 15l-4 1 1-4 9.5-9.5z"/></svg>',
    "trash": '<svg fill="none" stroke="currentColor" stroke-width="1.8" viewBox="0 0 24 24"><path d="M3 6h18M19 6v14a2 2 0 01-2 2H7a2 2 0 01-2-2V6m3 0V4a2 2 0 012-2h4a2 2 0 012 2v2"/></svg>',
    "wa": '<svg fill="currentColor" viewBox="0 0 24 24"><path d="M17.472 14.382c-.297-.149-1.758-.867-2.03-.967-.273-.099-.471-.148-.67.15-.197.297-.767.966-.94 1.164-.173.199-.347.223-.644.075-.297-.15-1.255-.463-2.39-1.475-.883-.788-1.48-1.761-1.653-2.059-.173-.297-.018-.458.13-.606.134-.133.298-.347.446-.52.149-.174.198-.298.298-.497.099-.198.05-.371-.025-.52-.075-.149-.669-1.612-.916-2.207-.242-.579-.487-.5-.669-.51-.173-.008-.371-.01-.57-.01-.198 0-.52.074-.792.372-.272.297-1.04 1.016-1.04 2.479 0 1.462 1.065 2.875 1.213 3.074.149.198 2.096 3.2 5.077 4.487.709.306 1.262.489 1.694.625.712.227 1.36.195 1.871.118.571-.085 1.758-.719 2.006-1.413.248-.694.248-1.289.173-1.413-.074-.124-.272-.198-.57-.347z"/></svg>',
    "whatsapp": '<svg fill="currentColor" viewBox="0 0 24 24"><path d="M17.472 14.382c-.297-.149-1.758-.867-2.03-.967-.273-.099-.471-.148-.67.15-.197.297-.767.966-.94 1.164-.173.199-.347.223-.644.075-.297-.15-1.255-.463-2.39-1.475-.883-.788-1.48-1.761-1.653-2.059-.173-.297-.018-.458.13-.606.134-.133.298-.347.446-.52.149-.174.198-.298.298-.497.099-.198.05-.371-.025-.52-.075-.149-.669-1.612-.916-2.207-.242-.579-.487-.5-.669-.51-.173-.008-.371-.01-.57-.01-.198 0-.52.074-.792.372-.272.297-1.04 1.016-1.04 2.479 0 1.462 1.065 2.875 1.213 3.074.149.198 2.096 3.2 5.077 4.487.709.306 1.262.489 1.694.625.712.227 1.36.195 1.871.118.571-.085 1.758-.719 2.006-1.413.248-.694.248-1.289.173-1.413-.074-.124-.272-.198-.57-.347z"/></svg>',
    "download": '<svg fill="none" stroke="currentColor" stroke-width="1.8" viewBox="0 0 24 24"><path d="M4 16v1a3 3 0 003 3h10a3 3 0 003-3v-1m-4-4l-4 4m0 0l-4-4m4 4V4"/></svg>',
    "eye": '<svg fill="none" stroke="currentColor" stroke-width="1.8" viewBox="0 0 24 24"><path d="M1 12s4-8 11-8 11 8 11 8-4 8-11 8-11-8-11-8z"/><circle cx="12" cy="12" r="3"/></svg>',
    "copy": '<svg fill="none" stroke="currentColor" stroke-width="1.8" viewBox="0 0 24 24"><rect x="9" y="9" width="13" height="13" rx="2" ry="2"/><path d="M5 15H4a2 2 0 01-2-2V4a2 2 0 012-2h9a2 2 0 012 2v1"/></svg>',
    "form_settings": '<svg fill="none" stroke="currentColor" stroke-width="1.8" viewBox="0 0 24 24"><path d="M12 15a3 3 0 100-6 3 3 0 000 6z"/><path d="M19.4 15a1.65 1.65 0 00.33 1.82l.06.06a2 2 0 010 2.83 2 2 0 01-2.83 0l-.06-.06a1.65 1.65 0 00-1.82-.33 1.65 1.65 0 00-1 1.51V21a2 2 0 01-2 2 2 2 0 01-2-2v-.09A1.65 1.65 0 009 19.4a1.65 1.65 0 00-1.82.33l-.06.06a2 2 0 01-2.83 0 2 2 0 010-2.83l.06-.06a1.65 1.65 0 00.33-1.82 1.65 1.65 0 00-1.51-1H3a2 2 0 01-2-2 2 2 0 012-2h.09A1.65 1.65 0 004.6 9a1.65 1.65 0 00-.33-1.82l-.06-.06a2 2 0 010-2.83 2 2 0 012.83 0l.06.06a1.65 1.65 0 001.82.33H9a1.65 1.65 0 001-1.51V3a2 2 0 012-2 2 2 0 012 2v.09a1.65 1.65 0 001 1.51 1.65 1.65 0 001.82-.33l.06-.06a2 2 0 012.83 0 2 2 0 010 2.83l-.06.06a1.65 1.65 0 00-.33 1.82V9a1.65 1.65 0 001.51 1H21a2 2 0 012 2 2 2 0 01-2 2h-.09a1.65 1.65 0 00-1.51 1z"/></svg>',
}

def icon(name, size=14):
    return f'<span style="display:inline-flex;width:{size}px;height:{size}px;flex-shrink:0">{ICON.get(name,"")}</span>'


class CRMHandler(http.server.BaseHTTPRequestHandler):
    def log_message(self, fmt, *args): pass

    # ---- Response helpers ----
    def send_html(self, title, body_content):
        html = f"""<!DOCTYPE html>
<html lang="es"><head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>{title} | POSGRADO CRM ESAM</title>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800;900&display=swap" rel="stylesheet">
<style>{BASE_CSS}</style>
</head><body>{body_content}</body></html>"""
        b = html.encode("utf-8")
        self.send_response(200)
        self.send_header("Content-Type", "text/html; charset=utf-8")
        self.send_header("Content-Length", str(len(b)))
        self.send_header("Connection", "close")
        self.end_headers()
        self.wfile.write(b)

    def redirect(self, loc, cookie=None):
        self.send_response(302)
        self.send_header("Location", loc)
        self.send_header("Content-Length", "0")
        self.send_header("Connection", "close")
        if cookie: self.send_header("Set-Cookie", cookie)
        self.end_headers()

    def parse_body(self):
        n = int(self.headers.get("Content-Length", 0))
        return urllib.parse.parse_qs(self.rfile.read(n).decode("utf-8"))

    def parse_body_and_files(self):
        ctype = self.headers.get("Content-Type", "")
        n = int(self.headers.get("Content-Length", 0))
        raw_bytes = self.rfile.read(n) if n > 0 else b""
        
        params = {}
        files = {}
        first_uploaded_filename = None
        
        if "multipart/form-data" in ctype and "boundary=" in ctype:
            boundary_str = ctype.split("boundary=")[1].strip()
            if boundary_str.startswith('"') and boundary_str.endswith('"'):
                boundary_str = boundary_str[1:-1]
            boundary = boundary_str.encode("utf-8")
            
            parts = raw_bytes.split(b"--" + boundary)
            import re
            for part in parts:
                if not part or part.startswith(b"--") or part == b"\r\n":
                    continue
                if b"\r\n\r\n" not in part:
                    continue
                header_raw, content_raw = part.split(b"\r\n\r\n", 1)
                if content_raw.endswith(b"\r\n"):
                    content_raw = content_raw[:-2]
                
                header_text = header_raw.decode("utf-8", errors="ignore")
                disp_line = ""
                for line in header_text.split("\r\n"):
                    if line.lower().startswith("content-disposition:"):
                        disp_line = line
                        break
                
                if "form-data" in disp_line:
                    m_name = re.search(r'name="([^"]+)"', disp_line)
                    if not m_name:
                        continue
                    field_name = m_name.group(1)
                    
                    m_file = re.search(r'filename="([^"]+)"', disp_line)
                    if m_file and m_file.group(1):
                        orig_fn = m_file.group(1)
                        if orig_fn.strip():
                            ext = os.path.splitext(orig_fn)[1].lower() or ".bin"
                            saved_fn = f"doc_{int(time.time()*1000)}_{random.randint(1000,9999)}{ext}"
                            saved_path = os.path.join("uploads", saved_fn)
                            with open(saved_path, "wb") as f_out:
                                f_out.write(content_raw)
                            saved_url = f"/uploads/{saved_fn}"
                            files[field_name] = saved_url
                            if not first_uploaded_filename:
                                first_uploaded_filename = saved_url
                    else:
                        if field_name not in params:
                            params[field_name] = []
                        params[field_name].append(content_raw.decode("utf-8", errors="ignore"))
        else:
            params = urllib.parse.parse_qs(raw_bytes.decode("utf-8", errors="ignore"))
            
        return params, first_uploaded_filename, files

    def parse_body(self):
        params, _, _ = self.parse_body_and_files()
        return params

    def get_user_session(self):
        uid = get_uid_from_cookie(self.headers.get("Cookie", ""))
        return get_user(uid) if uid else None

    # ---- Toast ----
    def toast(self, msg):
        if not msg: return ""
        is_ok = any(w in msg for w in ["exitosamente", "creado", "actualizado", "habilitado", "generado"])
        bg = "rgba(16,185,129,0.1)" if is_ok else "rgba(245,158,11,0.1)"
        bc = "rgba(16,185,129,0.2)" if is_ok else "rgba(245,158,11,0.2)"
        fc = "#34d399" if is_ok else "#fbbf24"
        return f'<div class="toast" id="t"><div style="background:{bg};border:1px solid {bc};color:{fc};padding:10px 16px;border-radius:10px;font-size:11px;font-weight:600;display:flex;align-items:center;gap:8px">{esc(msg)}<button onclick="this.parentElement.parentElement.remove()" style="background:none;border:none;color:inherit;cursor:pointer;font-size:14px;line-height:1;opacity:0.6">&times;</button></div></div><script>setTimeout(function(){{var e=document.getElementById("t");if(e)e.remove()}},4000)</script>'

    # ---- Shell (sidebar + header) ----
    def shell(self, tab, title, content, user):
        is_admin = user["role"] == "ADMIN"

        # Navigation
        main_tabs = [
            ("dashboard", "Dashboard", "/dashboard"),
            ("registrations", "Inscripciones" if is_admin else "Mis Contactos", "/registrations"),
            ("links", "Enlaces", "/links"),
        ]
        admin_tabs = [
            ("programs", "Programas", "/programs"),
            ("users", "Usuarios", "/users"),
            ("form_settings", "Config. Formulario", "/settings/form")
        ] if is_admin else []

        nav = '<div class="nav-section">Principal</div>'
        for key, label, href in main_tabs:
            ac = " active" if key == tab else ""
            nav += f'<a href="{href}" class="nav-item{ac}">{icon(key)}{label}</a>'
        if admin_tabs:
            nav += '<div class="nav-section" style="margin-top:6px">Administracion</div>'
            for key, label, href in admin_tabs:
                ac = " active" if key == tab else ""
                nav += f'<a href="{href}" class="nav-item{ac}">{icon(key)}{label}</a>'

        role_bg = "rgba(139,92,246,0.1)" if is_admin else "rgba(59,130,246,0.1)"
        role_color = "#a78bfa" if is_admin else "#60a5fa"
        role_border = "rgba(139,92,246,0.2)" if is_admin else "rgba(59,130,246,0.2)"
        role_label = "ADMIN" if is_admin else "ASESOR"
        initials = "".join(w[0] for w in user["name"].split()[:2]).upper()

        self.send_html(title, f"""
<div style="display:flex;min-height:100vh">
  <aside style="width:220px;background:#0a0f1a;border-right:1px solid #151d2e;display:flex;flex-direction:column;position:fixed;top:0;left:0;bottom:0;z-index:40">
    <div style="padding:16px 14px;border-bottom:1px solid #151d2e;display:flex;align-items:center;gap:10px">
      <div style="width:30px;height:30px;border-radius:8px;background:linear-gradient(135deg,#2563eb,#06b6d4);display:flex;align-items:center;justify-content:center;color:white;font-weight:900;font-size:11px;box-shadow:0 2px 8px rgba(37,99,235,0.25)">P</div>
      <div><p style="font-weight:700;color:white;font-size:12px;line-height:1">POSGRADO CRM ESAM</p><p style="font-size:9px;color:#334155;font-weight:500;margin-top:2px">Plataforma Comercial</p></div>
    </div>
    <nav style="flex:1;padding:8px;display:flex;flex-direction:column;gap:1px;overflow-y:auto">{nav}</nav>
    <div style="padding:8px;border-top:1px solid #151d2e">
      <div style="padding:10px;border-radius:8px;background:#0f172a;border:1px solid #1e293b;margin-bottom:6px;display:flex;align-items:center;gap:8px">
        <div style="width:28px;height:28px;border-radius:7px;background:{role_bg};border:1px solid {role_border};display:flex;align-items:center;justify-content:center;color:{role_color};font-size:9px;font-weight:800;flex-shrink:0">{initials}</div>
        <div style="min-width:0">
          <p style="font-weight:600;color:#e2e8f0;font-size:11px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis">{esc(user["name"])}</p>
          <p style="font-size:9px;color:#475569;white-space:nowrap;overflow:hidden;text-overflow:ellipsis">{esc(user["email"])}</p>
        </div>
      </div>
      <a href="/logout" style="display:block;text-align:center;padding:6px;border-radius:6px;background:#0f172a;border:1px solid #1e293b;color:#64748b;font-size:10px;font-weight:600;text-decoration:none;transition:all 0.15s" onmouseover="this.style.color='#f87171';this.style.borderColor='rgba(239,68,68,0.2)'" onmouseout="this.style.color='#64748b';this.style.borderColor='#1e293b'">Cerrar Sesion</a>
    </div>
  </aside>
  <div style="margin-left:220px;flex:1;display:flex;flex-direction:column;min-width:0">
    <header style="height:48px;border-bottom:1px solid #1e293b;background:rgba(10,15,26,0.9);backdrop-filter:blur(12px);padding:0 20px;display:flex;align-items:center;justify-content:space-between;position:sticky;top:0;z-index:30">
      <h2 style="font-size:13px;font-weight:700;color:white">{esc(title)}</h2>
      <div style="display:flex;align-items:center;gap:8px">
        <span class="badge" style="background:{role_bg};color:{role_color};border-color:{role_border}">{role_label}</span>
        <span style="font-size:11px;color:#94a3b8;font-weight:600">{esc(user["name"].split()[0])}</span>
      </div>
    </header>
    <main class="fade-in" style="padding:20px;flex:1">{content}</main>
  </div>
</div>""")

    # ============================================================
    # ROUTING
    # ============================================================
    def do_GET(self):
        p = urllib.parse.urlparse(self.path)
        path = p.path
        qs = urllib.parse.parse_qs(p.query)
        msg = qs.get("msg", [""])[0]

        # Public routes
        if path.startswith("/uploads/"):
            fname = path[len("/uploads/"):]
            fpath = os.path.join("uploads", fname)
            if os.path.isfile(fpath):
                with open(fpath, "rb") as f:
                    content = f.read()
                ext = fname.split(".")[-1].lower()
                mime_map = {
                    "png": "image/png",
                    "jpg": "image/jpeg",
                    "jpeg": "image/jpeg",
                    "webp": "image/webp",
                    "gif": "image/gif",
                    "pdf": "application/pdf",
                    "doc": "application/msword",
                    "docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
                }
                mime = mime_map.get(ext, "application/octet-stream")
                self.send_response(200)
                self.send_header("Content-Type", mime)
                if "download" in qs:
                    self.send_header("Content-Disposition", f'attachment; filename="{fname}"')
                self.send_header("Content-Length", str(len(content)))
                self.send_header("Connection", "close")
                self.end_headers()
                self.wfile.write(content)
                return
            self.send_response(404)
            self.end_headers()
            return

        if path in ("/", "/login"): return self.page_login()
        if path.startswith("/f/"): return self.page_public_form(path[3:])
        if path == "/logout":
            uid = get_uid_from_cookie(self.headers.get("Cookie", ""))
            if uid:
                for k, v in list(db["sessions"].items()):
                    if v == uid: db["sessions"].pop(k, None)
            return self.redirect("/login", "crm_session=; Max-Age=0; Path=/")

        # Auth required
        user = self.get_user_session()
        if not user: return self.redirect("/login")
        is_admin = user["role"] == "ADMIN"

        # Shared pages
        selected_month = qs.get("month", ["all"])[0]
        selected_advisor = qs.get("advisor", ["all"])[0]
        if path == "/dashboard": return self.page_dashboard(user, msg)
        if path == "/registrations": return self.page_registrations(user, msg, selected_month, selected_advisor)
        if path == "/links": return self.page_links(user, msg)

        if path.startswith("/export/registrations"):
            is_excel = "csv" not in path
            regs = db["registrations"] if is_admin else regs_for_advisor(user["id"])
            if selected_month != "all":
                regs = [r for r in regs if r.get("createdAt", "").startswith(selected_month)]
            if is_admin and selected_advisor != "all":
                regs = [r for r in regs if r.get("advisorId") == selected_advisor]

            parts = []
            if selected_month != "all": parts.append(selected_month)
            if is_admin and selected_advisor != "all": parts.append(user_name(selected_advisor).replace(" ", "_"))
            fn_suffix = "_".join(parts) if parts else "todas"

            if is_excel:
                host_hdr = self.headers.get("Host", "localhost:3000")
                host_url = f"http://{host_hdr}"
                b = generate_excel_bytes(regs, is_admin, selected_month, selected_advisor, host_url)
                fn = f"inscripciones_{fn_suffix}.xlsx"
                self.send_response(200)
                self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                self.send_header("Content-Disposition", f'attachment; filename="{fn}"')
                self.send_header("Content-Length", str(len(b)))
                self.end_headers()
                self.wfile.write(b)
                return
            else:
                import csv, io
                output = io.StringIO()
                writer = csv.writer(output)
                writer.writerow(["ID", "Fecha", "Nombres", "Apellidos", "Nombre Completo", "CI y Extension", "Email", "Telefono", "Direccion", "Ciudad", "Fecha Nacimiento", "Estado Civil", "Grado Academico", "Profesion", "Universidad", "CI Frente URL", "CI Reverso URL", "Programa", "Asesor", "Estado", "Notas"])
                for r in regs:
                    writer.writerow([
                        r.get("id",""), r.get("createdAt",""), r.get("firstNames",""), r.get("lastNames",""), r.get("fullName",""), r.get("ci",""),
                        r.get("email",""), r.get("phone",""), r.get("address",""), r.get("city",""), r.get("birthDate",""), r.get("civilStatus",""),
                        r.get("academicDegree",""), r.get("profession",""), r.get("university",""), r.get("ciFrontUrl",""), r.get("ciBackUrl",""),
                        program_name(r.get("programId")), user_name(r.get("advisorId")), r.get("status",""), r.get("notes","")
                    ])
                b = output.getvalue().encode("utf-8-sig")
                fn = f"inscripciones_{fn_suffix}.csv"
                self.send_response(200)
                self.send_header("Content-Type", "text/csv; charset=utf-8")
                self.send_header("Content-Disposition", f'attachment; filename="{fn}"')
                self.send_header("Content-Length", str(len(b)))
                self.end_headers()
                self.wfile.write(b)
                return

        # Registration status change (Available to both Admin and Advisor for their leads)
        if path.startswith("/registrations/status/"):
            parts = path.split("/registrations/status/")[1].split("/")
            if len(parts) == 2:
                rid, new_status = parts
                reg = next((r for r in db["registrations"] if r["id"] == rid), None)
                if reg and new_status in STATUS_PIPELINE:
                    if not is_admin and reg.get("advisorId") != user["id"]:
                        return self.redirect("/registrations?msg=Acceso+denegado")
                    reg["status"] = new_status
                    return self.redirect(f"/registrations?msg=Estado+actualizado+exitosamente")
            return self.redirect("/registrations")

        # Admin only
        if not is_admin: return self.redirect("/dashboard")

        if path == "/programs": return self.page_programs(user, msg)
        if path == "/programs/new": return self.page_program_form(user, None)
        if path.startswith("/programs/edit/"):
            prog = get_program(path.split("/programs/edit/")[1])
            return self.page_program_form(user, prog) if prog else self.redirect("/programs?msg=Programa+no+encontrado")
        if path.startswith("/programs/toggle/"):
            prog = get_program(path.split("/programs/toggle/")[1])
            if prog:
                prog["active"] = not prog["active"]
                return self.redirect(f"/programs?msg=Programa+{'habilitado' if prog['active'] else 'deshabilitado'}+exitosamente")
            return self.redirect("/programs?msg=No+encontrado")
        if path.startswith("/programs/delete/"):
            prog = get_program(path.split("/programs/delete/")[1])
            if prog:
                db["programs"].remove(prog)
                db["links"] = [l for l in db["links"] if l["programId"] != prog["id"]]
                return self.redirect("/programs?msg=Programa+eliminado+exitosamente")
            return self.redirect("/programs?msg=No+encontrado")

        if path == "/users": return self.page_users(user, msg)
        if path == "/users/new": return self.page_user_form(user, None)
        if path.startswith("/users/edit/"):
            t = get_user(path.split("/users/edit/")[1])
            return self.page_user_form(user, t) if t else self.redirect("/users?msg=Usuario+no+encontrado")
        if path.startswith("/users/toggle/"):
            t = get_user(path.split("/users/toggle/")[1])
            if t:
                if t["id"] == user["id"]: return self.redirect("/users?msg=No+puedes+desactivar+tu+cuenta")
                t["active"] = not t["active"]
                return self.redirect(f"/users?msg=Usuario+{'habilitado' if t['active'] else 'deshabilitado'}+exitosamente")
            return self.redirect("/users?msg=No+encontrado")
        if path.startswith("/users/delete/"):
            t = get_user(path.split("/users/delete/")[1])
            if t:
                if t["id"] == user["id"]: return self.redirect("/users?msg=No+puedes+eliminar+tu+cuenta")
                db["users"].remove(t)
                db["links"] = [l for l in db["links"] if l["advisorId"] != t["id"]]
                return self.redirect("/users?msg=Usuario+eliminado+exitosamente")
            return self.redirect("/users?msg=No+encontrado")

        if path == "/settings/form": return self.page_form_settings(user, msg)

        # Links management
        if path == "/links/new": return self.page_link_form(user)
        if path.startswith("/links/toggle/"):
            lid = path.split("/links/toggle/")[1]
            link = get_link(lid)
            if link:
                link["active"] = not link["active"]
                return self.redirect(f"/links?msg=Enlace+{'habilitado' if link['active'] else 'deshabilitado'}+exitosamente")
            return self.redirect("/links?msg=No+encontrado")
        if path.startswith("/links/delete/"):
            lid = path.split("/links/delete/")[1]
            link = get_link(lid)
            if link:
                db["links"].remove(link)
                return self.redirect("/links?msg=Enlace+eliminado+exitosamente")
            return self.redirect("/links?msg=No+encontrado")



        return self.redirect("/dashboard")

    def do_POST(self):
        p = urllib.parse.urlparse(self.path)
        path = p.path

        if path == "/login":
            params = self.parse_body()
            email = params.get("email", [""])[0].strip().lower()
            pw = params.get("password", [""])[0]
            u = next((x for x in db["users"] if x["email"] == email and x["password"] == pw and x["active"]), None)
            if u:
                tok = "sess_" + str(random.randint(100000, 999999))
                db["sessions"][tok] = u["id"]
                return self.redirect("/dashboard", f"crm_session={tok}; Path=/; HttpOnly; SameSite=Lax")
            return self.page_login("Credenciales invalidas o cuenta deshabilitada.")

        if path.startswith("/api/submit/"):
            code = path.split("/api/submit/")[1]
            link = get_link_by_code(code)
            if link and link["active"]:
                fs = db.get("form_settings", {
                    "datos_personales": True,
                    "documentos_ci": True,
                    "datos_contacto": True,
                    "datos_academicos": True
                })

                params, _, files = self.parse_body_and_files()
                firstNames = params.get("firstNames", [""])[0].strip()
                lastNames = params.get("lastNames", [""])[0].strip()
                fullName = f"{firstNames} {lastNames}".strip() if (firstNames or lastNames) else params.get("fullName", [""])[0].strip()
                ci = params.get("ci", [""])[0].strip()
                email = params.get("email", [""])[0].strip()
                address = params.get("address", [""])[0].strip()
                city = params.get("city", ["La Paz"])[0].strip()
                phone = params.get("phone", [""])[0].strip()
                birthDate = params.get("birthDate", [""])[0].strip()
                civilStatus = params.get("civilStatus", ["Soltero/a"])[0].strip()
                academicDegree = params.get("academicDegree", ["Licenciatura"])[0].strip()
                profession = params.get("profession", [""])[0].strip()
                university = params.get("university", [""])[0].strip()
                
                ciFrontUrl = files.get("ciFrontFile", "")
                ciBackUrl = files.get("ciBackFile", "")

                if fs.get("documentos_ci", True) and not (ciFrontUrl or ciBackUrl):
                    return self.send_html("Documento Requerido", """
<div style="min-height:100vh;display:flex;align-items:center;justify-content:center;padding:20px">
  <div class="card fade-in" style="padding:40px;text-align:center;max-width:440px;width:100%">
    <div style="width:64px;height:64px;border-radius:50%;background:rgba(239,68,68,0.12);border:1px solid rgba(239,68,68,0.25);display:flex;align-items:center;justify-content:center;margin:0 auto 16px;color:#f87171;font-size:28px;font-weight:900">⚠️</div>
    <h1 style="font-size:20px;font-weight:900;color:white;margin-bottom:8px">Archivo de C.I. Requerido</h1>
    <p style="font-size:12px;color:#94a3b8;line-height:1.6">Debes adjuntar al menos un archivo o fotografia de tu Cedula de Identidad para enviar la postulacion.</p>
    <button onclick="window.history.back()" class="btn btn-primary" style="margin-top:20px">Volver al Formulario</button>
  </div>
</div>""")

                if not fullName:
                    fullName = "Postulante Registrado"

                valid_submit = True
                if fs.get("datos_personales", True) and not (firstNames or fullName):
                    valid_submit = False
                if fs.get("datos_contacto", True) and not (email or phone):
                    valid_submit = False

                if valid_submit:

                    link["clickCount"] += 1
                    db["registrations"].insert(0, {
                        "id": gen_id("r"),
                        "fullName": fullName,
                        "firstNames": firstNames,
                        "lastNames": lastNames,
                        "ci": ci,
                        "email": email,
                        "address": address,
                        "city": city,
                        "phone": phone,
                        "birthDate": birthDate,
                        "civilStatus": civilStatus,
                        "ciFrontUrl": ciFrontUrl,
                        "ciBackUrl": ciBackUrl,
                        "academicDegree": academicDegree,
                        "profession": profession,
                        "university": university,
                        "programId": link["programId"],
                        "advisorId": link["advisorId"],
                        "linkId": link["id"],
                        "status": "NUEVO",
                        "createdAt": datetime.now().strftime("%Y-%m-%d %H:%M"),
                        "notes": ""
                    })
                    pname = program_name(link["programId"])
                    aname = user_name(link["advisorId"])
                    return self.send_html("Postulacion Enviada", f"""
<div style="min-height:100vh;display:flex;align-items:center;justify-content:center;padding:20px">
  <div class="card fade-in" style="padding:40px;text-align:center;max-width:460px;width:100%">
    <div style="width:64px;height:64px;border-radius:50%;background:rgba(16,185,129,0.12);border:1px solid rgba(16,185,129,0.25);display:flex;align-items:center;justify-content:center;margin:0 auto 16px;color:#34d399;font-size:28px;font-weight:900">✓</div>
    <h1 style="font-size:20px;font-weight:900;color:white;margin-bottom:8px">Postulacion Registrada con Exito</h1>
    <p style="font-size:12px;color:#94a3b8;line-height:1.6">Gracias <strong style="color:white">{esc(fullName)}</strong>, tu formulario y documentos han sido recepcionados correctamente.</p>
    <div style="margin:16px 0;padding:12px;border-radius:10px;background:rgba(59,130,246,0.08);border:1px solid rgba(59,130,246,0.18);text-align:left;font-size:11px">
      <p style="color:#60a5fa;font-weight:700">📌 Asesor Asignado: <span style="color:white">{esc(aname)}</span></p>
      <p style="color:#94a3b8;margin-top:4px">📚 Programa: <span style="color:white">{esc(pname)}</span></p>
    </div>
    <a href="/" class="btn btn-primary" style="margin-top:8px">Volver al inicio</a>
  </div>
</div>""")
                else:
                    return self.send_html("Campos Requeridos Faltantes", """
<div style="min-height:100vh;display:flex;align-items:center;justify-content:center;padding:20px">
  <div class="card fade-in" style="padding:40px;text-align:center;max-width:440px;width:100%">
    <div style="width:64px;height:64px;border-radius:50%;background:rgba(239,68,68,0.12);border:1px solid rgba(239,68,68,0.25);display:flex;align-items:center;justify-content:center;margin:0 auto 16px;color:#f87171;font-size:28px;font-weight:900">⚠️</div>
    <h1 style="font-size:20px;font-weight:900;color:white;margin-bottom:8px">Faltan Campos Obligatorios</h1>
    <p style="font-size:12px;color:#94a3b8;line-height:1.6">Por favor completa todos los campos requeridos marcados con asterisco (*) antes de enviar el formulario.</p>
    <button onclick="window.history.back()" class="btn btn-primary" style="margin-top:20px">Volver al Formulario</button>
  </div>
</div>""")
            return self.redirect("/")


        # Auth required
        user = self.get_user_session()
        if not user or user["role"] != "ADMIN": return self.redirect("/login")

        if path == "/programs/save":
            params, file_url, _ = self.parse_body_and_files()
            pid = params.get("id", [""])[0].strip()
            name = params.get("name", [""])[0].strip()
            code = params.get("code", [""])[0].strip().upper()
            ptype = params.get("type", ["CURSO"])[0]
            desc = params.get("description", [""])[0].strip()
            duration = params.get("duration", [""])[0].strip()
            modality = params.get("modality", [""])[0].strip()
            investment = params.get("investment", [""])[0].strip()
            input_url = params.get("imageUrl", [""])[0].strip()
            image_url = file_url or input_url

            if not name or not code: return self.redirect("/programs?msg=Nombre+y+codigo+obligatorios")
            if pid:
                prog = get_program(pid)
                if prog:
                    final_img = image_url if image_url else prog.get("imageUrl", "")
                    prog.update({"name": name, "code": code, "type": ptype, "description": desc, "duration": duration, "modality": modality, "investment": investment, "imageUrl": final_img})
                    return self.redirect("/programs?msg=Programa+actualizado+exitosamente")
                return self.redirect("/programs?msg=No+encontrado")
            
            new_id = gen_id("p")
            db["programs"].append({"id": new_id, "name": name, "code": code, "type": ptype, "active": True, "description": desc, "duration": duration, "modality": modality, "investment": investment, "imageUrl": image_url})
            
            # Generar de manera automatica enlaces de inscripcion para TODOS los asesores de ventas activos
            links_created = sync_advisor_links()

            return self.redirect(f"/programs?msg=Programa+creado+exitosamente+con+{links_created}+enlaces+generados+para+los+asesores")

        if path == "/users/save":
            params = self.parse_body()
            tid = params.get("id", [""])[0].strip()
            name = params.get("name", [""])[0].strip()
            email = params.get("email", [""])[0].strip().lower()
            pw = params.get("password", [""])[0].strip()
            role = params.get("role", ["ASESOR"])[0]
            phone = params.get("phone", [""])[0].strip()

            if not name or not email: return self.redirect("/users?msg=Nombre+y+email+obligatorios")
            if tid:
                t = get_user(tid)
                if t:
                    dup = next((u for u in db["users"] if u["email"] == email and u["id"] != tid), None)
                    if dup: return self.redirect(f"/users/edit/{tid}?msg=Email+ya+en+uso")
                    t.update({"name": name, "email": email, "role": role, "phone": phone})
                    if pw: t["password"] = pw
                    sync_advisor_links()
                    return self.redirect("/users?msg=Usuario+actualizado+exitosamente")
                return self.redirect("/users?msg=No+encontrado")
            dup = next((u for u in db["users"] if u["email"] == email), None)
            if dup: return self.redirect("/users/new?msg=Email+ya+en+uso")
            
            new_u_id = gen_id("u")
            db["users"].append({"id": new_u_id, "name": name, "email": email, "password": pw or "Asesor123!", "role": role, "phone": phone, "active": True, "createdAt": datetime.now().strftime("%Y-%m-%d %H:%M")})
            
            # Generar de manera automatica enlaces para todos los programas activos si es ASESOR
            links_created = sync_advisor_links()
                        
            return self.redirect(f"/users?msg=Usuario+creado+exitosamente+con+{links_created}+enlaces+asociados" if role == "ASESOR" else "/users?msg=Usuario+creado+exitosamente")

        if path == "/links/save":
            params = self.parse_body()
            pid = params.get("programId", [""])[0]
            aid = params.get("advisorId", [""])[0]
            if not pid or not aid: return self.redirect("/links?msg=Programa+y+asesor+obligatorios")
            # Check for duplicate
            existing = next((l for l in db["links"] if l["programId"] == pid and l["advisorId"] == aid), None)
            if existing: return self.redirect("/links?msg=Ya+existe+un+enlace+para+esa+combinacion")
            db["links"].append({"id": gen_id("l"), "code": gen_code(), "programId": pid, "advisorId": aid, "clickCount": 0, "active": True, "createdAt": datetime.now().strftime("%Y-%m-%d %H:%M")})
            return self.redirect("/links?msg=Enlace+generado+exitosamente")

        if path == "/settings/form/save":
            params = self.parse_body()
            db["form_settings"] = {
                "datos_personales": "datos_personales" in params,
                "documentos_ci": "documentos_ci" in params,
                "datos_contacto": "datos_contacto" in params,
                "datos_academicos": "datos_academicos" in params,
            }
            return self.redirect("/settings/form?msg=Configuracion+de+formulario+guardada+exitosamente")

        if path == "/registrations/notes":
            params = self.parse_body()
            rid = params.get("id", [""])[0]
            notes = params.get("notes", [""])[0].strip()
            reg = next((r for r in db["registrations"] if r["id"] == rid), None)
            if reg:
                reg["notes"] = notes
                return self.redirect("/registrations?msg=Notas+actualizadas+exitosamente")
            return self.redirect("/registrations")

        return self.redirect("/dashboard")

    # ============================================================
    # PAGES
    # ============================================================

    def page_login(self, error=""):
        err = f'<div style="padding:12px;border-radius:10px;background:rgba(239,68,68,0.1);border:1px solid rgba(239,68,68,0.2);color:#f87171;font-size:12px;font-weight:600;text-align:center;margin-bottom:16px">{esc(error)}</div>' if error else ""

        self.send_html("Iniciar Sesion", f"""
<div style="min-height:100vh;display:flex;align-items:center;justify-content:center;padding:20px">
  <div class="card fade-in" style="width:100%;max-width:420px;padding:32px">
    <div style="text-align:center;margin-bottom:24px">
      <div style="width:48px;height:48px;border-radius:14px;background:linear-gradient(135deg,#2563eb,#06b6d4);margin:0 auto 12px;display:flex;align-items:center;justify-content:center;color:white;font-weight:900;font-size:18px;box-shadow:0 8px 24px rgba(37,99,235,0.3)">P</div>
      <h1 style="font-size:22px;font-weight:900;color:white">POSGRADO CRM ESAM</h1>
      <p style="font-size:11px;color:#475569;margin-top:4px">Plataforma Comercial y Tracking de Inscripciones</p>
    </div>
    {err}
    <form method="POST" action="/login" style="display:flex;flex-direction:column;gap:14px">
      <div><label style="font-size:11px;font-weight:600;color:#94a3b8;display:block;margin-bottom:5px">Correo Electronico *</label><input type="email" name="email" required placeholder="correo@ejemplo.com" class="input"></div>
      <div><label style="font-size:11px;font-weight:600;color:#94a3b8;display:block;margin-bottom:5px">Contrasena *</label><input type="password" name="password" required placeholder="Ingresa tu contrasena" class="input"></div>
      <button type="submit" class="btn btn-primary" style="width:100%;justify-content:center;padding:12px;font-size:13px;margin-top:4px">Acceder al Sistema</button>
    </form>
  </div>
</div>""")

    # ---- DASHBOARD ----
    def page_dashboard(self, user, msg=""):
        is_admin = user["role"] == "ADMIN"
        regs = db["registrations"] if is_admin else regs_for_advisor(user["id"])
        links = db["links"] if is_admin else links_for_advisor(user["id"])

        curr_ym = datetime.now().strftime("%Y-%m")
        curr_month_label = format_ym_label(curr_ym)

        # Contactos recibidos en el mes en curso
        month_regs = [r for r in regs if r.get("createdAt", "").startswith(curr_ym)]

        total = len(regs)
        by_status = {s: len([r for r in regs if r["status"] == s]) for s in STATUS_PIPELINE}
        conv_rate = round(by_status["MATRICULADO"] / total * 100) if total > 0 else 0

        # Status summary cards
        status_cards = ""
        for s in STATUS_PIPELINE:
            cls, color = STATUS_COLORS[s]
            count = by_status[s]
            pct = round(count / total * 100) if total > 0 else 0
            status_cards += f"""<div class="card" style="padding:16px">
              <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:8px">
                <span style="font-size:10px;font-weight:600;color:#64748b">{STATUS_LABELS[s]}</span>
                <span class="badge {cls}">{count}</span>
              </div>
              <div class="progress-bar"><div class="progress-fill" style="width:{pct}%;background:var(--c,#3b82f6)"></div></div>
              <p style="font-size:10px;color:#475569;margin-top:4px">{pct}% del total</p>
            </div>"""

        # Table rows for current month contacts (or fallback to recent if none in current month yet)
        display_regs = month_regs if month_regs else regs
        rows = ""
        for r in display_regs[:6]:
            cls = STATUS_COLORS.get(r["status"], ("", ""))[0]
            rows += f"""<tr>
              <td style="font-weight:700;color:white">{esc(r["fullName"])}</td>
              <td style="color:#94a3b8">{esc(program_name(r["programId"]))}</td>
              <td style="color:#60a5fa;font-weight:600">{esc(user_name(r["advisorId"]))}</td>
              <td><span class="badge {cls}">{STATUS_LABELS.get(r["status"], r["status"])}</span></td>
              <td style="color:#475569;font-size:11px">{r["createdAt"]}</td>
            </tr>"""
        if not rows:
            rows = '<tr><td colspan="5" style="text-align:center;padding:32px;color:#475569;font-style:italic">Sin contactos recibidos este mes.</td></tr>'

        # Admin management cards
        admin_panel = ""
        if is_admin:
            admin_panel = f"""
<div style="display:grid;grid-template-columns:1fr 1fr;gap:16px">
  <a href="/programs" class="card card-lift" style="padding:20px;display:flex;align-items:center;gap:16px;text-decoration:none;color:inherit">
    <div style="width:48px;height:48px;border-radius:14px;background:rgba(139,92,246,0.12);border:1px solid rgba(139,92,246,0.2);display:flex;align-items:center;justify-content:center;color:#a78bfa;flex-shrink:0">{ICON["programs"]}</div>
    <div><p style="font-size:14px;font-weight:700;color:white">Gestion de Programas</p><p style="font-size:11px;color:#475569;margin-top:2px">{len(db["programs"])} programas | Crear, editar, eliminar</p></div>
  </a>
  <a href="/users" class="card card-lift" style="padding:20px;display:flex;align-items:center;gap:16px;text-decoration:none;color:inherit">
    <div style="width:48px;height:48px;border-radius:14px;background:rgba(6,182,212,0.12);border:1px solid rgba(6,182,212,0.2);display:flex;align-items:center;justify-content:center;color:#22d3ee;flex-shrink:0">{ICON["users"]}</div>
    <div><p style="font-size:14px;font-weight:700;color:white">Gestion de Usuarios</p><p style="font-size:11px;color:#475569;margin-top:2px">{len(db["users"])} usuarios | Crear, editar, eliminar</p></div>
  </a>
</div>"""

        content = f"""{self.toast(msg)}
<div style="display:flex;flex-direction:column;gap:20px">
  <div class="card" style="padding:24px;display:flex;justify-content:space-between;align-items:center">
    <div>
      <span class="badge" style="background:rgba(59,130,246,0.1);color:#60a5fa;border-color:rgba(59,130,246,0.2);margin-bottom:8px">CRM Comercial Enterprise</span>
      <h1 style="font-size:20px;font-weight:900;color:white;margin-top:6px">{"Panel Administrativo General" if is_admin else f"Mi Panel - {esc(user['name'])}"}</h1>
      <p style="font-size:12px;color:#475569;margin-top:4px">{"Metricas globales de todo el equipo comercial" if is_admin else "Tus metricas personales de captacion"}</p>
    </div>
    <div style="display:flex;gap:8px">
      <a href="/links" class="btn btn-primary">Ver Enlaces</a>
      <a href="/registrations?month={curr_ym}" class="btn btn-secondary">Ver Contactos del Mes</a>
    </div>
  </div>

  <!-- KPIs -->
  <div style="display:grid;grid-template-columns:repeat(5,1fr);gap:14px">
    <a href="/registrations?month={curr_ym}" class="card card-lift" style="padding:18px;border-color:rgba(59,130,246,0.4);background:linear-gradient(135deg,rgba(37,99,235,0.15),rgba(15,23,42,0.9));text-decoration:none;color:inherit">
      <div style="display:flex;justify-content:space-between;align-items:center">
        <p style="font-size:11px;font-weight:700;color:#60a5fa">Contactos del Mes</p>
        <span class="badge" style="background:rgba(59,130,246,0.2);color:#93c5fd;border-color:rgba(59,130,246,0.3);font-size:8px">{curr_month_label}</span>
      </div>
      <p style="font-size:32px;font-weight:900;color:white;margin-top:4px">{len(month_regs)}</p>
      <p style="font-size:10px;color:#94a3b8;margin-top:4px">Ver filtro del mes &rarr;</p>
    </a>
    <div class="card card-lift" style="padding:18px">
      <p style="font-size:11px;font-weight:600;color:#475569">Total Historico</p>
      <p style="font-size:32px;font-weight:900;color:white;margin-top:4px">{total}</p>
      <p style="font-size:10px;color:#475569;margin-top:4px">{"Todo el equipo" if is_admin else "Tus leads"}</p>
    </div>
    <div class="card card-lift" style="padding:18px">
      <p style="font-size:11px;font-weight:600;color:#475569">Enlaces Activos</p>
      <p style="font-size:32px;font-weight:900;color:#60a5fa;margin-top:4px">{len([l for l in links if l["active"]])}</p>
      <p style="font-size:10px;color:#475569;margin-top:4px">{sum(l["clickCount"] for l in links)} clics totales</p>
    </div>
    <div class="card card-lift" style="padding:18px">
      <p style="font-size:11px;font-weight:600;color:#475569">Matriculados</p>
      <p style="font-size:32px;font-weight:900;color:#34d399;margin-top:4px">{by_status["MATRICULADO"]}</p>
      <p style="font-size:10px;color:#475569;margin-top:4px">Conversion completada</p>
    </div>
    <div class="card card-lift" style="padding:18px">
      <p style="font-size:11px;font-weight:600;color:#475569">Tasa Conversion</p>
      <p style="font-size:32px;font-weight:900;color:#fbbf24;margin-top:4px">{conv_rate}%</p>
      <div class="progress-bar" style="margin-top:8px"><div class="progress-fill" style="width:{conv_rate}%;background:linear-gradient(90deg,#fbbf24,#f59e0b)"></div></div>
    </div>
  </div>

  <!-- Status Pipeline -->
  <div class="card" style="padding:20px">
    <h3 style="font-size:13px;font-weight:700;color:white;margin-bottom:16px">Pipeline de Conversion Total</h3>
    <div style="display:grid;grid-template-columns:repeat(5,1fr);gap:12px">{status_cards}</div>
  </div>

  {admin_panel}

  <!-- Recent contacts received in current month table -->
  <div class="card" style="padding:20px">
    <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:16px">
      <div>
        <h3 style="font-size:14px;font-weight:800;color:white">Contactos Recibidos en el Mes en Curso ({curr_month_label})</h3>
        <p style="font-size:11px;color:#475569;margin-top:2px">{len(month_regs)} contactos registrados durante {curr_month_label}</p>
      </div>
      <a href="/registrations?month={curr_ym}" class="btn btn-sm btn-primary">Ver Todos los de {curr_month_label} &rarr;</a>
    </div>
    <div style="border-radius:12px;border:1px solid #1e293b;overflow:hidden">
      <table><thead><tr><th>Postulante</th><th>Programa</th><th>Asesor</th><th>Estado</th><th>Fecha</th></tr></thead><tbody>{rows}</tbody></table>
    </div>
  </div>
</div>"""
        self.shell("dashboard", "Panel de Control", content, user)

    # ---- REGISTRATIONS ----
    def page_registrations(self, user, msg="", selected_month="all", selected_advisor="all"):
        is_admin = user["role"] == "ADMIN"
        regs = db["registrations"] if is_admin else regs_for_advisor(user["id"])

        curr_ym = datetime.now().strftime("%Y-%m")
        curr_month_name = format_ym_label(curr_ym)

        # Apply Month filter
        month_filtered = regs
        if selected_month != "all":
            month_filtered = [r for r in regs if r.get("createdAt", "").startswith(selected_month)]

        # Apply Advisor filter (Admin only)
        filtered_regs = month_filtered
        if is_admin and selected_advisor != "all":
            filtered_regs = [r for r in filtered_regs if r.get("advisorId") == selected_advisor]

        # Dropdown options listing all 12 calendar months dynamically
        month_options = get_all_12_months_options(selected_month, regs)

        # Admin Advisor Filter Dropdown and Metrics
        advisor_cards_markup = ""
        adv_options = ""
        if is_admin:
            advisors = [u for u in db["users"] if u.get("role") == "ASESOR"]
            cards_html = ""
            for adv in advisors:
                adv_id = adv["id"]
                # Count contacts generated by this advisor under current month filter if set
                adv_regs = [r for r in month_filtered if r.get("advisorId") == adv_id]
                adv_count = len(adv_regs)
                adv_matriculados = len([r for r in adv_regs if r.get("status") == "MATRICULADO"])
                
                sel = "selected" if selected_advisor == adv_id else ""
                adv_options += f'<option value="{adv_id}" {sel}>{esc(adv["name"])} — [{adv_count} contactos]</option>'

                is_active = selected_advisor == adv_id
                card_border = "border-color:rgba(139,92,246,0.5);background:linear-gradient(135deg,rgba(139,92,246,0.15),rgba(15,23,42,0.9))" if is_active else ""
                card_link = f"/registrations?month={selected_month}&advisor={adv_id}" if selected_month != "all" else f"/registrations?advisor={adv_id}"
                cards_html += f"""<a href="{card_link}" class="card card-lift" style="padding:14px;text-decoration:none;color:inherit;{card_border}">
                  <div style="display:flex;justify-content:space-between;align-items:center">
                    <p style="font-size:11px;font-weight:700;color:{'#a78bfa' if is_active else '#94a3b8'}">{esc(adv["name"])}</p>
                    <span class="badge" style="background:rgba(139,92,246,0.2);color:#c4b5fd;border-color:rgba(139,92,246,0.3);font-size:9px">{adv_count} contactos</span>
                  </div>
                  <p style="font-size:22px;font-weight:900;color:white;margin-top:4px">{adv_count}</p>
                  <p style="font-size:10px;color:#64748b;margin-top:2px">{adv_matriculados} matriculados</p>
                </a>"""

            advisor_cards_markup = f"""
<div style="display:flex;flex-direction:column;gap:6px;margin-bottom:8px">
  <div style="display:flex;justify-content:space-between;align-items:center">
    <span style="font-size:11px;font-weight:800;color:#a78bfa">👥 Rendimiento Comercial por Asesor de Ventas:</span>
    <a href="/registrations?advisor=all" style="font-size:10px;color:#94a3b8;font-weight:600;text-decoration:none">Ver Todos ({len(month_filtered)})</a>
  </div>
  <div style="display:grid;grid-template-columns:repeat({min(len(advisors), 4)}, 1fr);gap:12px">
    {cards_html}
  </div>
</div>"""
        else:
            # Breakdown of contacts by Program Link for this specific Advisor
            adv_links = links_for_advisor(user["id"])
            prog_cards_html = ""
            for l in adv_links:
                prog = get_program(l["programId"])
                pname = prog["name"] if prog else "Programa"
                pcode = prog["code"] if prog else "CODE"
                ptype = prog.get("type", "CURSO") if prog else "CURSO"
                
                # Count contacts from this advisor for this program/link
                link_regs = [r for r in regs if r.get("linkId") == l["id"] or r.get("programId") == l["programId"]]
                l_count = len(link_regs)
                l_matriculados = len([r for r in link_regs if r.get("status") == "MATRICULADO"])

                prog_cards_html += f"""<div class="card card-lift" style="padding:14px;background:rgba(15,23,42,0.9);border:1px solid rgba(59,130,246,0.25)">
                  <div style="display:flex;justify-content:space-between;align-items:center">
                    <span class="badge {TYPE_COLORS.get(ptype,'')}" style="font-size:8px">{esc(pcode)}</span>
                    <span style="font-size:10px;font-weight:700;color:#34d399">{l_matriculados} matriculados</span>
                  </div>
                  <p style="font-size:12px;font-weight:800;color:white;margin-top:6px;line-height:1.2;white-space:nowrap;overflow:hidden;text-overflow:ellipsis" title="{esc(pname)}">{esc(pname)}</p>
                  <div style="display:flex;justify-content:space-between;align-items:baseline;margin-top:8px">
                    <span style="font-size:20px;font-weight:900;color:#60a5fa">{l_count} <span style="font-size:10px;color:#94a3b8;font-weight:500">contactos</span></span>
                    <a href="/f/{l['code']}" target="_blank" style="font-size:10px;color:#38bdf8;text-decoration:none;font-weight:600">🔗 Ver Enlace</a>
                  </div>
                </div>"""

            advisor_cards_markup = f"""
<div style="display:flex;flex-direction:column;gap:6px;margin-bottom:8px">
  <div style="display:flex;justify-content:space-between;align-items:center">
    <span style="font-size:11px;font-weight:800;color:#60a5fa">📊 Mis Contactos Recibidos por Enlace de Programa:</span>
    <span style="font-size:10px;color:#94a3b8">Inscripciones captadas mediante tus enlaces únicos</span>
  </div>
  <div style="display:grid;grid-template-columns:repeat(auto-fill, minmax(220px, 1fr));gap:12px">
    {prog_cards_html}
  </div>
</div>"""

        # Table Rows
        rows = ""
        for r in filtered_regs:
            cls = STATUS_COLORS.get(r["status"], ("", ""))[0]
            pname = program_name(r["programId"])
            aname = user_name(r["advisorId"])
            wa = "591" + r.get("phone", "").replace(" ", "").replace("+", "")

            # Status change dropdown (enabled for both Admin and Advisor)
            opts = ""
            for s in STATUS_PIPELINE:
                sel = "selected" if s == r["status"] else ""
                opts += f'<option value="{s}" {sel}>{STATUS_LABELS[s]}</option>'
            status_select = f'<select class="input" style="font-size:10px;padding:4px 24px 4px 8px;width:auto;min-width:120px" onchange="window.location=\'/registrations/status/{r["id"]}/\'+this.value">{opts}</select>'

            notes_preview = esc(r.get("notes", "")[:30]) + ("..." if len(r.get("notes", "")) > 30 else "") if r.get("notes") else '<span style="color:#334155;font-style:italic">Sin notas</span>'

            ci_badge = ""
            c1 = r.get("ciFrontUrl", "")
            c2 = r.get("ciBackUrl", "")
            if c1 or c2:
                ci_badge = f'<span class="badge" style="background:rgba(139,92,246,0.15);color:#c4b5fd;border-color:rgba(139,92,246,0.3);font-size:9px" title="Fotos de C.I. adjuntadas">🪪 Adjuntado ({2 if c1 and c2 else 1})</span>'
            else:
                ci_badge = '<span style="font-size:10px;color:#475569;font-style:italic">Sin fotos</span>'

            # Build JSON payload for modal popup
            reg_json = json.dumps({
                "id": r.get("id", ""),
                "fullName": r.get("fullName", ""),
                "firstNames": r.get("firstNames", ""),
                "lastNames": r.get("lastNames", ""),
                "ci": r.get("ci", ""),
                "email": r.get("email", ""),
                "phone": r.get("phone", ""),
                "address": r.get("address", ""),
                "city": r.get("city", ""),
                "birthDate": r.get("birthDate", ""),
                "civilStatus": r.get("civilStatus", ""),
                "academicDegree": r.get("academicDegree", ""),
                "profession": r.get("profession", ""),
                "university": r.get("university", ""),
                "ciFrontUrl": r.get("ciFrontUrl", ""),
                "ciBackUrl": r.get("ciBackUrl", ""),
                "programName": pname,
                "advisorName": aname,
                "status": r.get("status", ""),
                "createdAt": r.get("createdAt", ""),
                "notes": r.get("notes", "")
            }).replace('"', '&quot;')

            rows += f"""<tr>
              <td style="color:#475569;font-size:11px;white-space:nowrap">{r["createdAt"]}</td>
              <td><p style="font-weight:700;color:white">{esc(r["fullName"])}</p><p style="font-size:10px;color:#475569;font-family:monospace">CI: {esc(r["ci"])}</p></td>
              <td><p style="color:#94a3b8">{esc(r["email"])}</p><p style="font-size:10px;color:#475569">Tel: {esc(r["phone"])} &middot; {esc(r.get("city",""))}</p></td>
              <td style="color:#94a3b8;font-weight:600;font-size:11px">{esc(pname)}</td>
              <td><p style="font-weight:700;color:#34d399;font-size:11px">{esc(r.get("academicDegree",""))}</p><p style="font-size:10px;color:#94a3b8">{esc(r.get("profession",""))}</p></td>
              <td style="color:#60a5fa;font-weight:700"><a href="/registrations?advisor={r['advisorId']}" style="color:#60a5fa;text-decoration:none" title="Filtrar inscripciones captadas por {esc(aname)}">{esc(aname)}</a></td>
              <td>{ci_badge}</td>
              <td>{status_select}</td>
              <td style="text-align:right;white-space:nowrap">
                <button onclick="openRegistrationModal({reg_json})" class="btn btn-sm btn-secondary" style="margin-right:4px">👁️ Detalle</button>
                <a href="https://wa.me/{wa}" target="_blank" class="btn btn-success btn-sm" title="WhatsApp">{ICON["whatsapp"]} WA</a>
              </td>
            </tr>"""
        if not rows:
            rows = f'<tr><td colspan="9" style="text-align:center;padding:40px;color:#475569;font-style:italic">No se encontraron contactos para los filtros seleccionados.</td></tr>'

        # Active Filters Banner
        active_badges = ""
        if selected_month != "all":
            active_badges += f'<span class="badge" style="background:rgba(59,130,246,0.2);color:#93c5fd;border-color:rgba(59,130,246,0.3);font-size:10px">📅 Mes: {format_ym_label(selected_month)} {"⭐ (Mes Actual)" if selected_month == curr_ym else ""}</span> '
        if is_admin and selected_advisor != "all":
            active_badges += f'<span class="badge" style="background:rgba(139,92,246,0.2);color:#c4b5fd;border-color:rgba(139,92,246,0.3);font-size:10px">👤 Asesor: {esc(user_name(selected_advisor))}</span> '

        active_filter_banner = f"""
<div style="display:flex;align-items:center;justify-content:space-between;padding:10px 14px;border-radius:10px;background:rgba(59,130,246,0.08);border:1px solid rgba(59,130,246,0.18)">
  <div style="display:flex;align-items:center;gap:8px">
    <span style="font-size:11px;color:#60a5fa;font-weight:700">Filtros Activos:</span>
    {active_badges}
    <a href="/registrations?month=all&advisor=all" style="font-size:10px;color:#f87171;font-weight:700;text-decoration:none;margin-left:8px">✖ Limpiar Filtros</a>
  </div>
  <span style="font-size:11px;font-weight:700;color:#34d399">{len(filtered_regs)} contactos encontrados</span>
</div>""" if active_badges else ""

        modal_markup = """
<div id="regDetailModal" style="display:none;position:fixed;inset:0;z-index:999;background:rgba(2,6,23,0.85);backdrop-filter:blur(8px);align-items:center;justify-content:center;padding:20px">
  <div class="card fade-in" style="max-width:720px;width:100%;max-height:90vh;overflow-y:auto;padding:28px;border:1px solid rgba(59,130,246,0.3);box-shadow:0 20px 50px rgba(0,0,0,0.6)">
    <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:20px;padding-bottom:12px;border-bottom:1px solid rgba(255,255,255,0.08)">
      <div>
        <h2 style="font-size:18px;font-weight:900;color:white" id="modalFullName">Detalle de la Postulacion</h2>
        <p style="font-size:11px;color:#60a5fa;margin-top:2px" id="modalProgramAdvisor"></p>
      </div>
      <button onclick="document.getElementById('regDetailModal').style.display='none'" class="btn btn-sm btn-secondary" style="font-size:18px;padding:2px 10px">&times;</button>
    </div>
    
    <div id="modalBodyContent" style="display:flex;flex-direction:column;gap:18px"></div>
  </div>
</div>

<script>
function renderDocCard(title, url) {
  if (!url) return '<div style="flex:1;padding:12px;border-radius:10px;background:rgba(30,41,59,0.3);border:1px dashed #334155"><p style="font-size:11px;font-weight:700;color:#64748b;margin-bottom:4px">'+title+'</p><p style="font-size:11px;color:#475569;font-style:italic">No adjuntado (Opcional)</p></div>';
  
  var isImg = url.match(/\.(png|jpg|jpeg|webp|gif|svg)$/i);
  var preview = isImg 
    ? '<a href="'+url+'" target="_blank"><img src="'+url+'" style="width:100%;height:140px;object-fit:cover;border-radius:8px;border:1px solid #334155;margin-bottom:8px"></a>'
    : '<div style="width:100%;height:140px;border-radius:8px;background:#020617;border:1px solid #334155;display:flex;flex-direction:column;align-items:center;justify-content:center;gap:6px;margin-bottom:8px"><span style="font-size:28px">📄</span><span style="font-size:10px;color:#94a3b8;font-weight:700">Documento Adjunto (PDF/Word)</span></div>';
    
  return '<div style="flex:1;padding:14px;border-radius:12px;background:rgba(15,23,42,0.9);border:1px solid rgba(139,92,246,0.3);display:flex;flex-direction:column;justify-content:space-between">' +
    '<div>' +
      '<p style="font-size:11px;font-weight:800;color:#a78bfa;margin-bottom:8px">'+title+'</p>' +
      preview +
    '</div>' +
    '<div style="display:flex;gap:6px;margin-top:6px">' +
      '<a href="'+url+'" target="_blank" class="btn btn-sm btn-secondary" style="flex:1;justify-content:center;font-size:10px">👁️ Ver</a>' +
      '<a href="'+url+'?download=1" download class="btn btn-sm btn-primary" style="flex:1;justify-content:center;font-size:10px">📥 Descargar</a>' +
    '</div>' +
  '</div>';
}

function openRegistrationModal(d) {
  document.getElementById('modalFullName').innerText = '📋 Postulante: ' + (d.fullName || (d.firstNames + ' ' + d.lastNames));
  document.getElementById('modalProgramAdvisor').innerText = '📚 Programa: ' + d.programName + ' | 👤 Asesor: ' + d.advisorName;
  
  var frontCard = renderDocCard('🪪 Anverso / Documento Principal (C.I.)', d.ciFrontUrl);
  var backCard = renderDocCard('🪪 Reverso / Documento Secundario (C.I.)', d.ciBackUrl);
  
  var html = '' +
    '<div style="padding:16px;border-radius:12px;background:rgba(15,23,42,0.8);border:1px solid rgba(59,130,246,0.2);display:flex;flex-direction:column;gap:10px">' +
      '<span style="font-size:12px;font-weight:900;color:#60a5fa;border-bottom:1px solid rgba(255,255,255,0.06);padding-bottom:6px">👤 DATOS PERSONALES</span>' +
      '<div style="display:grid;grid-template-columns:1fr 1fr;gap:10px;font-size:11px">' +
        '<div><strong style="color:#64748b">Nombres:</strong> <span style="color:white;font-weight:700">'+(d.firstNames||d.fullName)+'</span></div>' +
        '<div><strong style="color:#64748b">Apellidos:</strong> <span style="color:white;font-weight:700">'+(d.lastNames||'--')+'</span></div>' +
        '<div><strong style="color:#64748b">C.I. y Extension:</strong> <span style="color:#fbbf24;font-weight:700">'+(d.ci||'--')+'</span></div>' +
        '<div><strong style="color:#64748b">Correo electronico:</strong> <span style="color:#94a3b8">'+(d.email||'--')+'</span></div>' +
        '<div><strong style="color:#64748b">Celular / WhatsApp:</strong> <span style="color:#34d399;font-weight:700">'+(d.phone||'--')+'</span></div>' +
        '<div><strong style="color:#64748b">Ciudad:</strong> <span style="color:white">'+(d.city||'--')+'</span></div>' +
        '<div><strong style="color:#64748b">Direccion:</strong> <span style="color:#cbd5e1">'+(d.address||'--')+'</span></div>' +
        '<div><strong style="color:#64748b">Fecha de Nacimiento:</strong> <span style="color:white">'+(d.birthDate||'--')+'</span></div>' +
        '<div><strong style="color:#64748b">Estado Civil:</strong> <span style="color:white">'+(d.civilStatus||'--')+'</span></div>' +
        '<div><strong style="color:#64748b">Fecha de Registro:</strong> <span style="color:#94a3b8">'+(d.createdAt||'--')+'</span></div>' +
      '</div>' +
    '</div>' +
    
    '<div style="padding:16px;border-radius:12px;background:rgba(15,23,42,0.8);border:1px solid rgba(16,185,129,0.2);display:flex;flex-direction:column;gap:10px">' +
      '<span style="font-size:12px;font-weight:900;color:#34d399;border-bottom:1px solid rgba(255,255,255,0.06);padding-bottom:6px">🎓 DATOS ACADEMICOS</span>' +
      '<div style="display:grid;grid-template-columns:1fr 1fr;gap:10px;font-size:11px">' +
        '<div><strong style="color:#64748b">Grado Academico:</strong> <span style="color:#34d399;font-weight:700">'+(d.academicDegree||'--')+'</span></div>' +
        '<div><strong style="color:#64748b">Profesion:</strong> <span style="color:white;font-weight:700">'+(d.profession||'--')+'</span></div>' +
        '<div style="grid-column:span 2"><strong style="color:#64748b">Universidad de Egreso:</strong> <span style="color:#60a5fa;font-weight:700">'+(d.university||'--')+'</span></div>' +
      '</div>' +
    '</div>' +

    '<div style="padding:16px;border-radius:12px;background:rgba(15,23,42,0.8);border:1px solid rgba(139,92,246,0.2);display:flex;flex-direction:column;gap:10px">' +
      '<span style="font-size:12px;font-weight:900;color:#a78bfa;border-bottom:1px solid rgba(255,255,255,0.06);padding-bottom:6px">🪪 DOCUMENTOS DE IDENTIDAD C.I.</span>' +
      '<div style="display:grid;grid-template-columns:1fr 1fr;gap:16px;margin-top:4px">' + frontCard + backCard + '</div>' +
    '</div>' +

    '<div style="padding:16px;border-radius:12px;background:rgba(15,23,42,0.8);border:1px solid rgba(245,158,11,0.25);display:flex;flex-direction:column;gap:10px">' +
      '<span style="font-size:12px;font-weight:900;color:#fbbf24;border-bottom:1px solid rgba(255,255,255,0.06);padding-bottom:6px">📝 SEGUIMIENTO COMERCIAL Y NOTAS</span>' +
      '<form method="POST" action="/registrations/notes" style="display:flex;flex-direction:column;gap:8px">' +
        '<input type="hidden" name="id" value="' + d.id + '">' +
        '<textarea name="notes" placeholder="Escribe observaciones o notas de seguimiento (ej. Llamada realizada, solicita facilidades de pago, enviara titulo el viernes...)" class="input" style="min-height:70px;resize:vertical;font-size:11px">' + (d.notes || '') + '</textarea>' +
        '<div style="display:flex;justify-content:space-between;align-items:center">' +
          '<span style="font-size:10px;color:#64748b">Guarda observaciones para mantener un historial de seguimiento continuo del lead.</span>' +
          '<button type="submit" class="btn btn-sm btn-primary">💾 Guardar Notas de Seguimiento</button>' +
        '</div>' +
      '</form>' +
    '</div>';

  document.getElementById('modalBodyContent').innerHTML = html;
  document.getElementById('regDetailModal').style.display = 'flex';
}
</script>"""

        content = f"""{self.toast(msg)}
{modal_markup}
<div style="display:flex;flex-direction:column;gap:20px">
  <div class="card" style="padding:20px;display:flex;justify-content:space-between;align-items:center">
    <div>
      <h1 style="font-size:20px;font-weight:900;color:white">{"Todas las Inscripciones" if is_admin else "Mis Contactos Captados"}</h1>
      <p style="font-size:12px;color:#475569;margin-top:4px">{"Gestion completa de postulantes, filtros por mes y desglose por asesor" if is_admin else "Postulantes generados a traves de tus enlaces"}</p>
    </div>
    <div style="display:flex;gap:10px;align-items:center;flex-wrap:wrap">
      <a href="/export/registrations.xlsx?month={selected_month}&advisor={selected_advisor}" class="btn btn-success" style="font-weight:700;padding:8px 16px;box-shadow:0 4px 14px rgba(16,185,129,0.25)">📊 Exportar a Excel (.xlsx)</a>
      <a href="/export/registrations.csv?month={selected_month}&advisor={selected_advisor}" class="btn btn-secondary" style="font-size:11px">{ICON["download"]} CSV</a>
      <span class="badge" style="background:rgba(59,130,246,0.1);color:#60a5fa;border-color:rgba(59,130,246,0.2);font-size:12px;padding:6px 14px">{len(filtered_regs)} de {len(regs)} contactos</span>
    </div>
  </div>

  {advisor_cards_markup}

  <!-- UNIFIED MONTH & ADVISOR FILTER FORM -->
  <form method="GET" action="/registrations" class="card" style="padding:18px 24px;background:linear-gradient(135deg,rgba(15,23,42,0.95),rgba(30,41,59,0.7));border:1px solid rgba(59,130,246,0.35);display:flex;flex-direction:column;gap:14px">
    <div style="display:flex;justify-content:space-between;align-items:center">
      <span style="font-size:13px;font-weight:900;color:white;display:flex;align-items:center;gap:6px">⚡ Filtro Conjunto de Inscripciones (Multicriterio)</span>
      <div style="display:flex;gap:8px">
        <a href="/registrations?month={curr_ym}&advisor={selected_advisor}" class="btn btn-sm {'btn-primary' if selected_month == curr_ym else 'btn-secondary'}">⭐ Mes Actual ({curr_month_name})</a>
        <a href="/registrations?month=all&advisor=all" class="btn btn-sm {'btn-primary' if selected_month == 'all' and selected_advisor == 'all' else 'btn-secondary'}">📋 Ver Todo</a>
      </div>
    </div>

    <div style="display:grid;grid-template-columns:{'1fr 1fr' if is_admin else '1fr'};gap:16px;align-items:center">
      <div>
        <label style="font-size:11px;font-weight:700;color:#60a5fa;display:block;margin-bottom:6px">📅 Filtrar por Mes Calendario:</label>
        <select name="month" class="input" style="font-weight:700;font-size:12px;border-color:rgba(59,130,246,0.4);background:#020617;color:#60a5fa;cursor:pointer" onchange="this.form.submit()">
          <option value="all" {'selected' if selected_month == 'all' else ''}>🌐 Todos los Meses (Historico - {len(regs)} contactos)</option>
          {month_options}
        </select>
      </div>

      {f'''<div>
        <label style="font-size:11px;font-weight:700;color:#a78bfa;display:block;margin-bottom:6px">👤 Filtrar por Asesor de Ventas:</label>
        <select name="advisor" class="input" style="font-weight:700;font-size:12px;border-color:rgba(139,92,246,0.4);background:#020617;color:#c4b5fd;cursor:pointer" onchange="this.form.submit()">
          <option value="all" {'selected' if selected_advisor == 'all' else ''}>👥 Todos los Asesores de Ventas ({len(month_filtered)} contactos)</option>
          {adv_options}
        </select>
      </div>''' if is_admin else ''}
    </div>

    {active_filter_banner}
  </form>

  <div class="card" style="overflow:hidden">
    <table><thead><tr><th>Fecha</th><th>Postulante</th><th>Contacto / Ciudad</th><th>Programa</th><th>Grado / Profesion</th><th>Asesor</th><th>Fotos C.I.</th><th>Estado</th><th style="text-align:right">Acciones</th></tr></thead><tbody>{rows}</tbody></table>
  </div>
</div>"""
        self.shell("registrations", "Gestion de Inscripciones", content, user)

    # ---- LINKS ----
    def page_links(self, user, msg=""):
        sync_advisor_links()
        is_admin = user["role"] == "ADMIN"
        # Admin y Asesor ven unicamente SUS PROPIOS enlaces personales por programa
        links = links_for_advisor(user["id"])

        cards = ""
        for l in links:
            prog = get_program(l["programId"])
            pname = prog["name"] if prog else "Programa eliminado"
            pcode = prog["code"] if prog else "N/A"
            ptype = prog.get("type", "CURSO") if prog else "CURSO"
            duration = prog.get("duration", "") if prog else ""
            modality = prog.get("modality", "") if prog else ""

            reg_count = len(regs_for_link(l["id"]))
            active = l["active"]
            status_dot = '<span style="color:#34d399;font-size:9px;font-weight:700">● Activo</span>' if active else '<span style="color:#f87171;font-size:9px;font-weight:700">● Inactivo</span>'
            url = f"/f/{l['code']}"

            # Pre-filled WhatsApp message URL
            wa_text = urllib.parse.quote(f"¡Hola! Te comparto la información y el formulario de postulación para *{pname}* ({pcode}): http://localhost:3000{url}")
            wa_share_url = f"https://wa.me/?text={wa_text}"

            cards += f"""<div class="card card-lift" style="padding:16px;display:flex;flex-direction:column;justify-content:space-between;border:1px solid #1e293b;background:#0f172a;border-radius:12px">
              <div>
                <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:8px">
                  <span class="badge {TYPE_COLORS.get(ptype,'')}" style="font-size:8px;padding:2px 6px">{esc(pcode)} &middot; {esc(ptype)}</span>
                  {status_dot}
                </div>
                <h3 style="font-size:13px;font-weight:800;color:white;line-height:1.35;margin-bottom:4px">{esc(pname)}</h3>
                <p style="font-size:10px;color:#64748b;margin-bottom:10px;display:flex;align-items:center;gap:6px"><span>⏱️ {esc(duration or 'Varias duraciones')}</span> &middot; <span>📍 {esc(modality or 'Virtual/Presencial')}</span></p>

                <div style="padding:8px 10px;border-radius:8px;background:#020617;border:1px solid #1e293b;display:flex;flex-direction:column;gap:4px">
                  <span style="font-size:9px;font-weight:700;color:#475569;text-transform:uppercase;letter-spacing:0.5px">Enlace Unico Rastreable</span>
                  <div style="display:flex;align-items:center;justify-content:space-between;gap:6px">
                    <span style="font-family:monospace;font-size:10px;color:#60a5fa;overflow:hidden;text-overflow:ellipsis;white-space:nowrap">http://localhost:3000{url}</span>
                  </div>
                </div>
              </div>

              <div style="margin-top:12px;padding-top:10px;border-top:1px solid #1e293b;display:flex;flex-direction:column;gap:8px">
                <div style="display:flex;justify-content:space-between;align-items:center;font-size:10px">
                  <span style="color:#64748b"><strong style="color:#60a5fa">{l["clickCount"]}</strong> clics</span>
                  <a href="/registrations?month=all" style="color:#34d399;font-weight:700;text-decoration:none;font-size:9px" title="Ver postulantes que ingresaron por este enlace">
                    <span class="badge" style="background:rgba(16,185,129,0.12);color:#34d399;border-color:rgba(16,185,129,0.25);font-size:9px;padding:2px 7px">Postulantes: {reg_count}</span>
                  </a>
                </div>

                <div style="display:grid;grid-template-columns:1fr 1fr;gap:6px">
                  <button onclick="copyToClipboard('http://localhost:3000{url}', '{esc(pname)}')" class="btn btn-sm btn-primary" style="justify-content:center;padding:5px 8px;font-size:10px;gap:4px">{icon("copy", 11)} Copiar</button>
                  <a href="{wa_share_url}" target="_blank" class="btn btn-sm btn-success" style="justify-content:center;padding:5px 8px;font-size:10px;gap:4px">{icon("whatsapp", 11)} WhatsApp</a>
                </div>

                <a href="{url}" target="_blank" class="btn btn-sm btn-secondary" style="justify-content:center;font-size:10px;padding:5px 8px;gap:4px">{icon("eye", 11)} Probar Formulario Publico</a>
              </div>
            </div>"""

        if not cards:
            cards = '<div style="grid-column:span 3;text-align:center;padding:40px;color:#475569" class="card">No hay enlaces disponibles.</div>'

        content = f"""{self.toast(msg)}
<div style="display:flex;flex-direction:column;gap:16px">
  <div class="card" style="padding:16px 20px;display:flex;justify-content:space-between;align-items:center">
    <div>
      <h1 style="font-size:18px;font-weight:900;color:white">Mis Enlaces Promocionales por Programa</h1>
      <p style="font-size:11px;color:#475569;margin-top:2px">Cada programa posee tu enlace unico de rastreo. Comparte tus enlaces con prospectos para captar postulaciones.</p>
    </div>
    <div style="display:flex;gap:8px;align-items:center">
      <span class="badge" style="background:rgba(59,130,246,0.1);color:#60a5fa;border-color:rgba(59,130,246,0.2);font-size:10px;padding:4px 10px">{len(links)} enlaces personales</span>
    </div>
  </div>

  <div style="display:grid;grid-template-columns:repeat(3,1fr);gap:14px">{cards}</div>
</div>

<script>
function copyToClipboard(text, programName) {{
  navigator.clipboard.writeText(text).then(function() {{
    alert("Enlace copiado al portapapeles: " + programName + " - " + text);
  }}).catch(function(err) {{
    prompt("Copia este enlace manualmente:", text);
  }});
}}
</script>"""
        self.shell("links", "Enlaces Rastreables", content, user)

    # ---- LINK FORM (admin) ----
    def page_link_form(self, user):
        prog_opts = "".join(f'<option value="{p["id"]}">{esc(p["name"])} ({p["code"]})</option>' for p in db["programs"] if p["active"])
        adv_opts = "".join(f'<option value="{u["id"]}">{esc(u["name"])} ({u["email"]})</option>' for u in db["users"] if u["active"] and u["role"] == "ASESOR")

        content = f"""
<div style="display:flex;flex-direction:column;gap:20px">
  <div style="display:flex;align-items:center;gap:10px">
    <a href="/links" class="btn btn-secondary">{ICON["back"]} Volver</a>
    <h1 style="font-size:20px;font-weight:900;color:white">Generar Nuevo Enlace Rastreable</h1>
  </div>
  <div style="max-width:600px">
    <form method="POST" action="/links/save" class="card" style="padding:24px;display:flex;flex-direction:column;gap:16px">
      <div>
        <label style="font-size:11px;font-weight:600;color:#94a3b8;display:block;margin-bottom:6px">Programa Academico *</label>
        <select name="programId" required class="input">{prog_opts}</select>
      </div>
      <div>
        <label style="font-size:11px;font-weight:600;color:#94a3b8;display:block;margin-bottom:6px">Asesor Asignado *</label>
        <select name="advisorId" required class="input">{adv_opts}</select>
      </div>
      <p style="font-size:11px;color:#475569;padding:12px;border-radius:8px;background:rgba(59,130,246,0.05);border:1px solid rgba(59,130,246,0.1)">Se generara un codigo unico de rastreo. Los postulantes que se registren a traves de este enlace seran asignados automaticamente al asesor seleccionado.</p>
      <div style="display:flex;gap:10px;margin-top:4px">
        <button type="submit" class="btn btn-primary">Generar Enlace</button>
        <a href="/links" class="btn btn-secondary">Cancelar</a>
      </div>
    </form>
  </div>
</div>"""
        self.shell("links", "Nuevo Enlace", content, user)

    # ---- PROGRAMS ----
    def page_programs(self, user, msg=""):
        rows = ""
        for p in db["programs"]:
            tcls = TYPE_COLORS.get(p["type"], "")
            active_badge = '<span class="badge" style="background:rgba(16,185,129,0.1);color:#34d399;border-color:rgba(16,185,129,0.2)">Activo</span>' if p["active"] else '<span class="badge" style="background:rgba(239,68,68,0.1);color:#f87171;border-color:rgba(239,68,68,0.2)">Inactivo</span>'
            toggle_cls = "btn-warning" if p["active"] else "btn-success"
            toggle_label = "Deshabilitar" if p["active"] else "Habilitar"
            lcount = len(links_for_program(p["id"]))
            rcount = len(regs_for_program(p["id"]))
            img_tag = f'<img src="{esc(p["imageUrl"])}" style="width:44px;height:44px;border-radius:8px;object-fit:cover;border:1px solid #1e293b;flex-shrink:0">' if p.get("imageUrl") else '<div style="width:44px;height:44px;border-radius:8px;background:rgba(59,130,246,0.1);border:1px solid rgba(59,130,246,0.2);display:flex;align-items:center;justify-content:center;color:#60a5fa;font-size:10px;font-weight:800;flex-shrink:0">IMG</div>'

            rows += f"""<tr>
              <td>
                <div style="display:flex;align-items:center;gap:12px">
                  {img_tag}
                  <div>
                    <p style="font-weight:700;color:white">{esc(p["name"])}</p>
                    <p style="font-size:10px;color:#475569;margin-top:2px">{esc(p.get("description","")[:60])}</p>
                  </div>
                </div>
              </td>
              <td><span class="badge {tcls}">{p["type"]}</span></td>
              <td style="font-family:monospace;color:#60a5fa;font-weight:700">{esc(p["code"])}</td>
              <td style="text-align:center;font-weight:600;color:#94a3b8">{lcount}</td>
              <td style="text-align:center;font-weight:600;color:#94a3b8">{rcount}</td>
              <td>{active_badge}</td>
              <td style="text-align:right;white-space:nowrap">
                <div style="display:flex;gap:4px;justify-content:flex-end">
                  <a href="/programs/edit/{p["id"]}" class="btn btn-sm btn-secondary">{ICON["edit"]} Editar</a>
                  <a href="/programs/toggle/{p["id"]}" class="btn btn-sm {toggle_cls}">{toggle_label}</a>
                  <a href="/programs/delete/{p["id"]}" onclick="return confirm('Eliminar este programa y sus enlaces asociados?')" class="btn btn-sm btn-danger">{ICON["trash"]} Eliminar</a>
                </div>
              </td>
            </tr>"""
        if not rows:
            rows = '<tr><td colspan="7" style="text-align:center;padding:40px;color:#475569;font-style:italic">No hay programas.</td></tr>'

        content = f"""{self.toast(msg)}
<div style="display:flex;flex-direction:column;gap:20px">
  <div class="card" style="padding:20px;display:flex;justify-content:space-between;align-items:center">
    <div>
      <h1 style="font-size:20px;font-weight:900;color:white">Gestion de Programas Academicos</h1>
      <p style="font-size:12px;color:#475569;margin-top:4px">Crear, editar, habilitar/deshabilitar y eliminar programas</p>
    </div>
    <a href="/programs/new" class="btn btn-primary">{ICON["plus"]} Nuevo Programa</a>
  </div>
  <div class="card" style="overflow:hidden">
    <table><thead><tr><th>Programa</th><th>Tipo</th><th>Codigo</th><th style="text-align:center">Enlaces</th><th style="text-align:center">Inscritos</th><th>Estado</th><th style="text-align:right">Acciones</th></tr></thead><tbody>{rows}</tbody></table>
  </div>
</div>"""
        self.shell("programs", "Programas Academicos", content, user)

    def page_program_form(self, user, prog):
        is_edit = prog is not None
        title = "Editar Programa" if is_edit else "Nuevo Programa"
        vals = prog if is_edit else {"id": "", "name": "", "code": "", "type": "CURSO", "description": "", "duration": "", "modality": "Virtual", "investment": "", "imageUrl": ""}
        sel = lambda f, v: "selected" if vals.get(f) == v else ""

        content = f"""
<div style="display:flex;flex-direction:column;gap:20px">
  <div style="display:flex;align-items:center;gap:10px">
    <a href="/programs" class="btn btn-secondary">{ICON["back"]} Volver</a>
    <h1 style="font-size:20px;font-weight:900;color:white">{title}</h1>
  </div>
  <div style="max-width:700px">
    <form method="POST" action="/programs/save" enctype="multipart/form-data" class="card" style="padding:24px;display:flex;flex-direction:column;gap:16px">
      <input type="hidden" name="id" value="{vals['id']}">
      <div>
        <label style="font-size:11px;font-weight:600;color:#94a3b8;display:block;margin-bottom:6px">Nombre del Programa *</label>
        <input type="text" name="name" value="{esc(vals['name'])}" required placeholder="Ej. Maestria en Educacion Superior" class="input">
      </div>
      <div style="display:grid;grid-template-columns:1fr 1fr;gap:12px">
        <div>
          <label style="font-size:11px;font-weight:600;color:#94a3b8;display:block;margin-bottom:6px">Codigo *</label>
          <input type="text" name="code" value="{esc(vals['code'])}" required placeholder="MED-2026" class="input" style="text-transform:uppercase;font-family:monospace">
        </div>
        <div>
          <label style="font-size:11px;font-weight:600;color:#94a3b8;display:block;margin-bottom:6px">Tipo *</label>
          <select name="type" class="input"><option value="MAESTRIA" {sel("type","MAESTRIA")}>Maestria</option><option value="ESPECIALIDAD" {sel("type","ESPECIALIDAD")}>Especialidad</option><option value="DIPLOMADO" {sel("type","DIPLOMADO")}>Diplomado</option><option value="CURSO" {sel("type","CURSO")}>Curso</option></select>
        </div>
      </div>
      <div>
        <label style="font-size:11px;font-weight:600;color:#94a3b8;display:block;margin-bottom:6px">Descripcion</label>
        <input type="text" name="description" value="{esc(vals.get('description',''))}" placeholder="Breve descripcion del programa" class="input">
      </div>
      <div style="padding:14px;border-radius:10px;background:rgba(30,41,59,0.3);border:1px dashed #334155;display:flex;flex-direction:column;gap:10px">
        <div>
          <label style="font-size:11px;font-weight:700;color:#60a5fa;display:block;margin-bottom:4px">📁 Subir Imagen desde mi Equipo (PNG, JPG, WEBP)</label>
          <input type="file" name="imageFile" accept="image/png, image/jpeg, image/jpg, image/webp" class="input" style="padding:6px;background:#0f172a;cursor:pointer">
          <p style="font-size:10px;color:#64748b;margin-top:4px">Selecciona un archivo local para que se guarde y replique automáticamente en los formularios públicos.</p>
        </div>
        <div style="display:flex;align-items:center;gap:8px;margin:2px 0">
          <div style="flex:1;height:1px;background:#1e293b"></div>
          <span style="font-size:9px;color:#475569;font-weight:700">O TAMBIEN</span>
          <div style="flex:1;height:1px;background:#1e293b"></div>
        </div>
        <div>
          <label style="font-size:11px;font-weight:600;color:#94a3b8;display:block;margin-bottom:6px">URL Externa o Presets</label>
          <input type="text" id="img_input" name="imageUrl" value="{esc(vals.get('imageUrl',''))}" placeholder="https://... o /uploads/marketing_ia.png" class="input">
          <div style="display:flex;gap:6px;margin-top:6px;align-items:center">
            <span style="font-size:10px;color:#475569">Presets:</span>
            <button type="button" onclick="document.getElementById('img_input').value='/uploads/marketing_ia.png'" class="btn btn-sm btn-secondary">Marketing IA</button>
            <button type="button" onclick="document.getElementById('img_input').value='/uploads/educacion_superior.png'" class="btn btn-sm btn-secondary">Educacion Superior</button>
          </div>
        </div>
        {f'<div style="margin-top:4px;display:flex;align-items:center;gap:10px;padding:8px;border-radius:8px;background:#0f172a"><span style="font-size:10px;color:#94a3b8;font-weight:600">Imagen actual:</span><img src="{esc(vals.get("imageUrl"))}" style="width:70px;height:42px;object-fit:cover;border-radius:6px;border:1px solid #1e293b"></div>' if vals.get("imageUrl") else ''}
      </div>
      <div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:12px">
        <div>
          <label style="font-size:11px;font-weight:600;color:#94a3b8;display:block;margin-bottom:6px">Duracion</label>
          <input type="text" name="duration" value="{esc(vals.get('duration',''))}" placeholder="Ej. 24 meses" class="input">
        </div>
        <div>
          <label style="font-size:11px;font-weight:600;color:#94a3b8;display:block;margin-bottom:6px">Modalidad</label>
          <select name="modality" class="input"><option value="Virtual" {sel("modality","Virtual")}>Virtual</option><option value="Presencial" {sel("modality","Presencial")}>Presencial</option><option value="Semipresencial" {sel("modality","Semipresencial")}>Semipresencial</option></select>
        </div>
        <div>
          <label style="font-size:11px;font-weight:600;color:#94a3b8;display:block;margin-bottom:6px">Inversion</label>
          <input type="text" name="investment" value="{esc(vals.get('investment',''))}" placeholder="Ej. 18.000 Bs" class="input">
        </div>
      </div>
      <div style="display:flex;gap:10px;margin-top:8px">
        <button type="submit" class="btn btn-primary">{"Guardar Cambios" if is_edit else "Crear Programa"}</button>
        <a href="/programs" class="btn btn-secondary">Cancelar</a>
      </div>
    </form>
  </div>
</div>"""
        self.shell("programs", title, content, user)

    # ---- USERS ----
    def page_users(self, user, msg=""):
        rows = ""
        for u in db["users"]:
            role_cls = "background:rgba(139,92,246,0.1);color:#a78bfa;border-color:rgba(139,92,246,0.2)" if u["role"] == "ADMIN" else "background:rgba(59,130,246,0.1);color:#60a5fa;border-color:rgba(59,130,246,0.2)"
            active_badge = '<span class="badge" style="background:rgba(16,185,129,0.1);color:#34d399;border-color:rgba(16,185,129,0.2)">Activo</span>' if u["active"] else '<span class="badge" style="background:rgba(239,68,68,0.1);color:#f87171;border-color:rgba(239,68,68,0.2)">Inactivo</span>'
            is_self = u["id"] == user["id"]
            self_tag = ' <span style="font-size:9px;color:#fbbf24;font-weight:700">(TU)</span>' if is_self else ""
            lcount = len(links_for_advisor(u["id"]))
            rcount = len(regs_for_advisor(u["id"]))

            if is_self:
                actions = f'<a href="/users/edit/{u["id"]}" class="btn btn-sm btn-secondary">{ICON["edit"]} Editar</a><span class="btn btn-sm" style="color:#334155;cursor:not-allowed;background:#0f172a;border:1px solid #1e293b">Protegido</span>'
            else:
                toggle_cls = "btn-warning" if u["active"] else "btn-success"
                toggle_label = "Deshabilitar" if u["active"] else "Habilitar"
                actions = f"""<a href="/users/edit/{u["id"]}" class="btn btn-sm btn-secondary">{ICON["edit"]} Editar</a>
                  <a href="/users/toggle/{u["id"]}" class="btn btn-sm {toggle_cls}">{toggle_label}</a>
                  <a href="/users/delete/{u["id"]}" onclick="return confirm('Eliminar este usuario y sus enlaces?')" class="btn btn-sm btn-danger">{ICON["trash"]}</a>"""

            rows += f"""<tr>
              <td><a href="/registrations?advisor={u['id']}" style="font-weight:700;color:white;text-decoration:none" title="Ver inscripciones de este asesor">{esc(u["name"])}{self_tag}</a></td>
              <td style="color:#94a3b8">{esc(u["email"])}</td>
              <td style="color:#475569">{esc(u.get("phone","N/A"))}</td>
              <td><span class="badge" style="{role_cls}">{u["role"]}</span></td>
              <td style="text-align:center;font-weight:600;color:#94a3b8">{lcount}</td>
              <td style="text-align:center;font-weight:600;color:#94a3b8"><a href="/registrations?advisor={u['id']}" class="badge" style="background:rgba(59,130,246,0.15);color:#60a5fa;border-color:rgba(59,130,246,0.3);text-decoration:none" title="Ver inscripciones captadas por {esc(u['name'])}">{rcount} contactos</a></td>
              <td>{active_badge}</td>
              <td style="text-align:right"><div style="display:flex;gap:4px;justify-content:flex-end">{actions}</div></td>
            </tr>"""

        content = f"""{self.toast(msg)}
<div style="display:flex;flex-direction:column;gap:20px">
  <div class="card" style="padding:20px;display:flex;justify-content:space-between;align-items:center">
    <div>
      <h1 style="font-size:20px;font-weight:900;color:white">Gestion de Usuarios</h1>
      <p style="font-size:12px;color:#475569;margin-top:4px">Crear, editar, habilitar/deshabilitar y eliminar cuentas</p>
    </div>
    <a href="/users/new" class="btn btn-primary">{ICON["plus"]} Nuevo Usuario</a>
  </div>
  <div class="card" style="overflow:hidden">
    <table><thead><tr><th>Nombre</th><th>Email</th><th>Telefono</th><th>Rol</th><th style="text-align:center">Enlaces</th><th style="text-align:center">Contactos</th><th>Estado</th><th style="text-align:right">Acciones</th></tr></thead><tbody>{rows}</tbody></table>
  </div>
</div>"""
        self.shell("users", "Usuarios del Sistema", content, user)

    def page_user_form(self, user, target):
        is_edit = target is not None
        title = "Editar Usuario" if is_edit else "Nuevo Usuario"
        v = target if is_edit else {"id": "", "name": "", "email": "", "phone": "", "role": "ASESOR"}
        sel = lambda val: "selected" if v.get("role") == val else ""
        pw_ph = "Dejar vacio para mantener actual" if is_edit else "Minimo 6 caracteres"
        pw_req = "" if is_edit else "required"

        content = f"""
<div style="display:flex;flex-direction:column;gap:20px">
  <div style="display:flex;align-items:center;gap:10px">
    <a href="/users" class="btn btn-secondary">{ICON["back"]} Volver</a>
    <h1 style="font-size:20px;font-weight:900;color:white">{title}</h1>
  </div>
  <div style="max-width:600px">
    <form method="POST" action="/users/save" class="card" style="padding:24px;display:flex;flex-direction:column;gap:16px">
      <input type="hidden" name="id" value="{v['id']}">
      <div>
        <label style="font-size:11px;font-weight:600;color:#94a3b8;display:block;margin-bottom:6px">Nombre Completo *</label>
        <input type="text" name="name" value="{esc(v['name'])}" required placeholder="Ej. Juan Perez Martinez" class="input">
      </div>
      <div style="display:grid;grid-template-columns:1fr 1fr;gap:12px">
        <div>
          <label style="font-size:11px;font-weight:600;color:#94a3b8;display:block;margin-bottom:6px">Email *</label>
          <input type="email" name="email" value="{esc(v['email'])}" required placeholder="usuario@posgrado.com" class="input">
        </div>
        <div>
          <label style="font-size:11px;font-weight:600;color:#94a3b8;display:block;margin-bottom:6px">Telefono</label>
          <input type="text" name="phone" value="{esc(v.get('phone',''))}" placeholder="+591 70000000" class="input">
        </div>
      </div>
      <div style="display:grid;grid-template-columns:1fr 1fr;gap:12px">
        <div>
          <label style="font-size:11px;font-weight:600;color:#94a3b8;display:block;margin-bottom:6px">Contrasena {"" if is_edit else "*"}</label>
          <input type="password" name="password" {pw_req} placeholder="{pw_ph}" class="input">
        </div>
        <div>
          <label style="font-size:11px;font-weight:600;color:#94a3b8;display:block;margin-bottom:6px">Rol *</label>
          <select name="role" class="input"><option value="ADMIN" {sel("ADMIN")}>Administrador</option><option value="ASESOR" {sel("ASESOR")}>Asesor de Ventas</option></select>
        </div>
      </div>
      <div style="display:flex;gap:10px;margin-top:8px">
        <button type="submit" class="btn btn-primary">{"Guardar Cambios" if is_edit else "Crear Usuario"}</button>
        <a href="/users" class="btn btn-secondary">Cancelar</a>
      </div>
    </form>
  </div>
</div>"""
        self.shell("users", title, content, user)

    # ---- FORM SETTINGS ----
    def page_form_settings(self, user, msg=""):
        fs = db.get("form_settings", {
            "datos_personales": True,
            "documentos_ci": True,
            "datos_contacto": True,
            "datos_academicos": True
        })

        sections = [
            ("datos_personales", "1. Datos Personales de Identidad", "Incluye Nombres, Apellidos, Cédula de Identidad, Fecha de Nacimiento y Estado Civil."),
            ("documentos_ci", "2. Documentación y Fotografías C.I.", "Incluye los archivos o fotografías adjuntas del carnet de identidad (Anverso y Reverso)."),
            ("datos_contacto", "3. Datos de Contacto y Residencia", "Incluye Correo Electrónico, Teléfono Celular, Dirección y Ciudad."),
            ("datos_academicos", "4. Perfil Académico y Experiencia Laboral", "Incluye Grado Académico, Profesión u Ocupación y Universidad de Egreso."),
        ]

        items_html = ""
        for key, label, desc in sections:
            is_req = fs.get(key, True)
            checked = "checked" if is_req else ""
            badge_style = "background:rgba(59,130,246,0.15);color:#60a5fa;border-color:rgba(59,130,246,0.3)" if is_req else "background:rgba(148,163,184,0.15);color:#94a3b8;border-color:rgba(148,163,184,0.3)"
            badge_text = "Obligatorio" if is_req else "Opcional"

            items_html += f"""
            <div style="padding:18px;border-radius:14px;background:#0f172a;border:1px solid #1e293b;display:flex;align-items:center;justify-content:space-between;gap:16px;margin-bottom:12px">
              <div>
                <div style="display:flex;align-items:center;gap:10px">
                  <h3 style="font-size:14px;font-weight:700;color:white">{esc(label)}</h3>
                  <span class="badge" style="{badge_style}">{badge_text}</span>
                </div>
                <p style="font-size:11px;color:#94a3b8;margin-top:4px">{esc(desc)}</p>
              </div>
              <label style="display:inline-flex;align-items:center;gap:8px;cursor:pointer;background:#020617;padding:8px 14px;border-radius:10px;border:1px solid #1e293b">
                <input type="checkbox" name="{key}" value="1" {checked} style="width:16px;height:16px;accent-color:#2563eb;cursor:pointer">
                <span style="font-size:11px;font-weight:700;color:{'#60a5fa' if is_req else '#94a3b8'}">{'Obligatorio' if is_req else 'Opcional'}</span>
              </label>
            </div>
            """

        content = f"""
        <div style="max-width:800px;margin:0 auto" class="fade-in">
          <div style="margin-bottom:20px;display:flex;align-items:center;justify-content:space-between">
            <div>
              <h1 style="font-size:22px;font-weight:900;color:white">Configuración del Formulario Público</h1>
              <p style="font-size:12px;color:#94a3b8;margin-top:4px">Define qué apartados a completar son obligatorios u opcionales para los estudiantes.</p>
            </div>
            <span class="badge" style="background:rgba(139,92,246,0.15);color:#c4b5fd;border-color:rgba(139,92,246,0.3)">Solo Administrador</span>
          </div>

          {self.toast(msg) if msg else ''}

          <form method="POST" action="/settings/form/save">
            <div class="card" style="padding:24px;border-radius:20px">
              <p style="font-size:11px;font-weight:700;color:#64748b;text-transform:uppercase;letter-spacing:1px;margin-bottom:16px">Carácter de Obligatoriedad por Apartado</p>
              
              {items_html}

              <div style="margin-top:20px;display:flex;justify-content:flex-end">
                <button type="submit" class="btn btn-primary" style="padding:10px 24px;font-size:12px;font-weight:700">💾 Guardar Configuración</button>
              </div>
            </div>
          </form>
        </div>
        """
        self.shell("form_settings", "Configuración del Formulario", content, user)

    # ---- PUBLIC FORM ----
    def page_public_form(self, code):
        link = get_link_by_code(code)
        if not link:
            return self.send_html("No encontrado", '<div style="min-height:100vh;display:flex;align-items:center;justify-content:center"><div class="card" style="padding:40px;text-align:center;max-width:400px"><h1 style="font-size:20px;font-weight:900;color:white;margin-bottom:8px">Enlace no valido</h1><p style="font-size:12px;color:#475569">Este enlace no existe.</p><a href="/" class="btn btn-primary" style="margin-top:16px">Inicio</a></div></div>')
        if not link["active"]:
            return self.send_html("Desactivado", '<div style="min-height:100vh;display:flex;align-items:center;justify-content:center"><div class="card" style="padding:40px;text-align:center;max-width:400px"><h1 style="font-size:20px;font-weight:900;color:white;margin-bottom:8px">Enlace desactivado</h1><p style="font-size:12px;color:#475569">Este enlace ha sido desactivado temporalmente.</p><a href="/" class="btn btn-primary" style="margin-top:16px">Inicio</a></div></div>')

        fs = db.get("form_settings", {
            "datos_personales": True,
            "documentos_ci": True,
            "datos_contacto": True,
            "datos_academicos": True
        })
        req_dp = "required" if fs.get("datos_personales", True) else ""
        req_doc = "required" if fs.get("documentos_ci", True) else ""
        req_dc = "required" if fs.get("datos_contacto", True) else ""
        req_da = "required" if fs.get("datos_academicos", True) else ""

        badge_dp = '<span class="badge" style="background:rgba(59,130,246,0.2);color:#93c5fd;font-size:9px">Obligatorio</span>' if fs.get("datos_personales", True) else '<span class="badge" style="background:rgba(148,163,184,0.15);color:#94a3b8;font-size:9px">Opcional</span>'
        badge_doc = '<span class="badge" style="background:rgba(139,92,246,0.2);color:#c4b5fd;font-size:9px">Obligatorio</span>' if fs.get("documentos_ci", True) else '<span class="badge" style="background:rgba(148,163,184,0.15);color:#94a3b8;font-size:9px">Opcional</span>'
        badge_dc = '<span class="badge" style="background:rgba(59,130,246,0.2);color:#93c5fd;font-size:9px">Obligatorio</span>' if fs.get("datos_contacto", True) else '<span class="badge" style="background:rgba(148,163,184,0.15);color:#94a3b8;font-size:9px">Opcional</span>'
        badge_da = '<span class="badge" style="background:rgba(16,185,129,0.2);color:#6ee7b7;font-size:9px">Obligatorio</span>' if fs.get("datos_academicos", True) else '<span class="badge" style="background:rgba(148,163,184,0.15);color:#94a3b8;font-size:9px">Opcional</span>'

        pname = program_name(link["programId"])
        pcode = program_code(link["programId"])
        aname = user_name(link["advisorId"])
        prog = get_program(link["programId"])
        desc = prog.get("description", "") if prog else ""
        duration = prog.get("duration", "") if prog else ""
        modality = prog.get("modality", "") if prog else ""
        investment = prog.get("investment", "") if prog else ""
        image_url = prog.get("imageUrl", "") if prog else ""

        desc_html = f'<p style="font-size:12px;color:#94a3b8;line-height:1.5;margin-top:4px">{esc(desc)}</p>' if desc else ''
        duration_badge = f'<span class="badge" style="background:rgba(59,130,246,0.1);color:#60a5fa;border-color:rgba(59,130,246,0.15)">⏱️ {esc(duration)}</span>' if duration else ''
        modality_badge = f'<span class="badge" style="background:rgba(16,185,129,0.1);color:#34d399;border-color:rgba(16,185,129,0.15)">📍 {esc(modality)}</span>' if modality else ''
        investment_badge = f'<span class="badge" style="background:rgba(245,158,11,0.1);color:#fbbf24;border-color:rgba(245,158,11,0.15)">💰 {esc(investment)}</span>' if investment else ''

        # Left 50%: Program Title Header FIRST, followed by Program Image Poster
        if image_url:
            img_poster_markup = f"""<div style="position:relative;width:100%;background:#0b0f19;border-top:1px solid #1e293b">
              <img src="{esc(image_url)}" style="width:100%;height:auto;max-height:680px;object-fit:cover;display:block" alt="{esc(pname)}">
            </div>"""
        else:
            img_poster_markup = f"""<div style="padding:40px 20px;background:linear-gradient(135deg,#0b0f19,#1e1b4b);border-top:1px solid #1e293b;text-align:center;color:#64748b;font-size:11px;font-style:italic">
              <span>🖼️ Afiche promocional del programa</span>
            </div>"""

        left_column_html = f"""<div style="border-radius:20px;overflow:hidden;border:1px solid #1e293b;background:#0f172a;box-shadow:0 20px 40px rgba(0,0,0,0.6);display:flex;flex-direction:column">
          <!-- Program Title Header (FIRST) -->
          <div style="padding:22px 24px;background:linear-gradient(135deg,#0f172a,#1e1b4b);display:flex;flex-direction:column;gap:8px">
            <div style="display:flex;align-items:center;gap:8px;flex-wrap:wrap">
              <span class="badge" style="background:rgba(59,130,246,0.2);color:#93c5fd;border-color:rgba(59,130,246,0.3);font-size:10px">{esc(pcode)}</span>
              <span class="badge" style="background:rgba(16,185,129,0.15);color:#34d399;border-color:rgba(16,185,129,0.3);font-size:9px">{esc(modality or 'Modalidad Virtual/Presencial')}</span>
            </div>
            <h2 style="font-size:22px;font-weight:900;color:white;line-height:1.25;letter-spacing:-0.3px">{esc(pname)}</h2>
            {desc_html}
            <div style="display:flex;gap:8px;flex-wrap:wrap;margin-top:6px">
              {duration_badge}
              {investment_badge}
            </div>
          </div>
          <!-- Program Image Poster (SECOND) -->
          {img_poster_markup}
        </div>"""

        self.send_html(f"Inscripcion - {pname}", f"""
<div class="public-form-wrapper" style="min-height:100vh;padding:40px 20px;display:flex;align-items:center;justify-content:center">
  <div style="max-width:1180px;width:100%;display:flex;flex-direction:column;gap:24px" class="fade-in">
    
    <!-- HEADER ARRIBA -->
    <div class="public-form-header" style="text-align:center;padding:12px;margin-bottom:4px">
      <div style="width:48px;height:48px;border-radius:14px;background:linear-gradient(135deg,#2563eb,#06b6d4);margin:0 auto 12px;display:flex;align-items:center;justify-content:center;color:white;font-weight:900;font-size:18px;box-shadow:0 8px 24px rgba(37,99,235,0.3)">P</div>
      <h1 style="font-size:26px;font-weight:900;color:white;letter-spacing:-0.5px">Formulario Oficial de Inscripcion</h1>
      <p style="font-size:13px;color:#94a3b8;margin-top:6px;font-weight:500">Completa tus datos personales y academicos para registrar tu postulacion al programa.</p>
      <div style="display:inline-flex;align-items:center;gap:8px;margin-top:12px;padding:6px 16px;border-radius:100px;background:rgba(59,130,246,0.1);border:1px solid rgba(59,130,246,0.25);box-shadow:0 4px 12px rgba(0,0,0,0.15)">
        <span style="font-size:13px">👤</span>
        <span style="font-size:11px;color:#60a5fa;font-weight:700">Asesor Comercial Asignado: {esc(aname)}</span>
      </div>
    </div>

    <!-- 50/50 SPLIT GRID -->
    <div class="public-form-grid">
      
      <!-- LEFT 50%: PROGRAM TITLE FIRST, THEN PROGRAM IMAGE -->
      <div class="public-form-poster">
        {left_column_html}
      </div>

      <!-- RIGHT 50%: RECOLECCION DE DATOS (FORM FIELDS) -->
      <div class="card public-form-card" style="padding:28px;border-radius:20px">
        <form method="POST" action="/api/submit/{code}" enctype="multipart/form-data" style="display:flex;flex-direction:column;gap:18px">
          
          <!-- DATOS PERSONALES -->
          <div style="padding:16px;border-radius:12px;background:rgba(15,23,42,0.8);border:1px solid rgba(59,130,246,0.25);display:flex;flex-direction:column;gap:12px">
            <div style="display:flex;align-items:center;justify-content:space-between;padding-bottom:8px;border-bottom:1px solid rgba(255,255,255,0.06)">
              <span style="font-size:12px;font-weight:900;color:#60a5fa">📋 DATOS PERSONALES</span>
              {badge_dp}
            </div>

            <div class="form-grid-2col">
              <div>
                <label style="font-size:10px;font-weight:700;color:#94a3b8;display:block;margin-bottom:4px">Nombres {'*' if req_dp else ''}</label>
                <input name="firstNames" type="text" {req_dp} placeholder="Ej. Roberto Carlos" class="input">
              </div>
              <div>
                <label style="font-size:10px;font-weight:700;color:#94a3b8;display:block;margin-bottom:4px">Apellidos {'*' if req_dp else ''}</label>
                <input name="lastNames" type="text" {req_dp} placeholder="Ej. Vargas Morales" class="input">
              </div>
            </div>

            <div class="form-grid-2col">
              <div>
                <label style="font-size:10px;font-weight:700;color:#94a3b8;display:block;margin-bottom:4px">C.I. y Extension {'*' if req_dp else ''}</label>
                <input name="ci" type="text" {req_dp} placeholder="Ej. 7654321 LP" class="input">
              </div>
              <div>
                <label style="font-size:10px;font-weight:700;color:#94a3b8;display:block;margin-bottom:4px">Correo electronico {'*' if req_dc else ''}</label>
                <input name="email" type="email" {req_dc} placeholder="correo@ejemplo.com" class="input">
              </div>
            </div>

            <div class="form-grid-2to1">
              <div>
                <label style="font-size:10px;font-weight:700;color:#94a3b8;display:block;margin-bottom:4px">Direccion {'*' if req_dc else ''}</label>
                <input name="address" type="text" {req_dc} placeholder="Ej. Av. 6 de Agosto #2450" class="input">
              </div>
              <div>
                <label style="font-size:10px;font-weight:700;color:#94a3b8;display:block;margin-bottom:4px">Ciudad {'*' if req_dc else ''}</label>
                <select name="city" {req_dc} class="input">
                  <option value="La Paz">La Paz</option>
                  <option value="Santa Cruz">Santa Cruz</option>
                  <option value="Cochabamba">Cochabamba</option>
                  <option value="Sucre">Sucre</option>
                  <option value="Oruro">Oruro</option>
                  <option value="Potosi">Potosi</option>
                  <option value="Tarija">Tarija</option>
                  <option value="Beni">Beni</option>
                  <option value="Pando">Pando</option>
                  <option value="Otra">Otra</option>
                </select>
              </div>
            </div>

            <div class="form-grid-3col">
              <div>
                <label style="font-size:10px;font-weight:700;color:#94a3b8;display:block;margin-bottom:4px">Celular / WhatsApp {'*' if req_dc else ''}</label>
                <input name="phone" type="tel" {req_dc} placeholder="77889900" class="input">
              </div>
              <div>
                <label style="font-size:10px;font-weight:700;color:#94a3b8;display:block;margin-bottom:4px">Nacimiento {'*' if req_dp else ''}</label>
                <input name="birthDate" type="date" {req_dp} class="input" style="color-scheme:dark;padding:7px 8px;font-size:11px">
              </div>
              <div>
                <label style="font-size:10px;font-weight:700;color:#94a3b8;display:block;margin-bottom:4px">Estado Civil {'*' if req_dp else ''}</label>
                <select name="civilStatus" {req_dp} class="input" style="font-size:11px">
                  <option value="Soltero/a">Soltero/a</option>
                  <option value="Casado/a">Casado/a</option>
                  <option value="Divorciado/a">Divorciado/a</option>
                  <option value="Viudo/a">Viudo/a</option>
                  <option value="Conviviente">Union Libre</option>
                </select>
              </div>
            </div>

            <!-- DOCUMENTOS C.I. -->
            <div style="margin-top:2px;padding:12px;border-radius:10px;background:rgba(30,41,59,0.4);border:1px dashed rgba(139,92,246,0.35);display:flex;flex-direction:column;gap:8px">
              <div style="display:flex;align-items:center;justify-content:space-between">
                <span style="font-size:11px;font-weight:800;color:#a78bfa">🪪 DOCUMENTACION DE C.I.</span>
                {badge_doc}
              </div>
              <p style="font-size:9px;color:#64748b">Adjunta fotografia o PDF de tu C.I.</p>
              <div class="form-grid-2col">
                <div>
                  <label style="font-size:9px;font-weight:700;color:#cbd5e1;display:block;margin-bottom:3px">Anverso (C.I.) {'*' if req_doc else ''}</label>
                  <input type="file" name="ciFrontFile" id="ciFrontFile" accept="image/*,application/pdf,.doc,.docx" {req_doc} class="input" style="padding:4px;font-size:10px;background:#0f172a;cursor:pointer">
                </div>
                <div>
                  <label style="font-size:9px;font-weight:700;color:#cbd5e1;display:block;margin-bottom:3px">Reverso (Opcional)</label>
                  <input type="file" name="ciBackFile" id="ciBackFile" accept="image/*,application/pdf,.doc,.docx" class="input" style="padding:4px;font-size:10px;background:#0f172a;cursor:pointer">
                </div>
              </div>
            </div>
          </div>

          <!-- DATOS ACADEMICOS -->
          <div style="padding:16px;border-radius:12px;background:rgba(15,23,42,0.8);border:1px solid rgba(16,185,129,0.25);display:flex;flex-direction:column;gap:12px">
            <div style="display:flex;align-items:center;justify-content:space-between;padding-bottom:8px;border-bottom:1px solid rgba(255,255,255,0.06)">
              <span style="font-size:12px;font-weight:900;color:#34d399">🎓 DATOS ACADEMICOS</span>
              {badge_da}
            </div>

            <div class="form-grid-2col">
              <div>
                <label style="font-size:10px;font-weight:700;color:#94a3b8;display:block;margin-bottom:4px">Grado Academico {'*' if req_da else ''}</label>
                <select name="academicDegree" {req_da} class="input">
                  <option value="Licenciatura">Licenciatura</option>
                  <option value="Tecnico Superior">Tecnico Superior</option>
                  <option value="Diplomado">Diplomado</option>
                  <option value="Especialidad">Especialidad</option>
                  <option value="Maestria">Maestria</option>
                  <option value="Doctorado">Doctorado</option>
                  <option value="Estudiante Universitario">Estudiante Universitario</option>
                </select>
              </div>
              <div>
                <label style="font-size:10px;font-weight:700;color:#94a3b8;display:block;margin-bottom:4px">Profesion {'*' if req_da else ''}</label>
                <input name="profession" type="text" {req_da} placeholder="Ej. Ingeniero Comercial" class="input">
              </div>
            </div>

            <div>
              <label style="font-size:10px;font-weight:700;color:#94a3b8;display:block;margin-bottom:4px">Universidad de Egreso {'*' if req_da else ''}</label>
              <input name="university" type="text" {req_da} placeholder="Ej. Universidad Mayor de San Andres (UMSA)" class="input">
            </div>
          </div>

          <button type="submit" class="btn btn-primary" style="width:100%;justify-content:center;padding:12px;font-size:13px;font-weight:800;margin-top:2px">✨ Enviar Formulario de Postulacion</button>

        </form>
      </div>

    </div>
  </div>
</div>""")


import socketserver, sys

class ThreadedHTTPServer(socketserver.ThreadingMixIn, http.server.HTTPServer):
    daemon_threads = True
    allow_reuse_address = True

if __name__ == "__main__":
    server = ThreadedHTTPServer(("0.0.0.0", PORT), CRMHandler)
    print(f"Posgrado CRM Server v3.0 (Multithreaded) running on http://localhost:{PORT}")
    sys.stdout.flush()
    server.serve_forever()
